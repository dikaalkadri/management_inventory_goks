import os
import io
import pandas as pd
import zipfile
from datetime import datetime
import re
import openpyxl

from modules.eresto_mass.cleaner import clean_mass_data
from modules.eresto.excel_generator import generate_excel

# Create dummy raw dataframe
data = {
    'Order Date': ['01/07/2026 10:00:00', '02/07/2026 12:00:00'],
    'Outlet': ['01. Taplau', '01. Taplau'],
    'Product': ['MIE GORENG GOKS', 'NASI GORENG GOKS'],
    'Quantity': [2, 3]
}
for i in range(15):
    data[f'Dummy_{i}'] = ['x', 'y']
    
df_raw = pd.DataFrame(data)

# Save to temp excel
df_raw.to_excel('temp_raw.xlsx', index=False)

# Run cleaner
cleaned = clean_mass_data('temp_raw.xlsx')
df = cleaned['df']

# Run generator
wb = openpyxl.Workbook()
master_products = ['MIE GORENG GOKS', 'NASI GORENG GOKS']
output_buffer = io.BytesIO()
generate_excel(
    df=df,
    date_col=cleaned['date_col'],
    product_col=cleaned['product_col'],
    qty_col=cleaned['qty_col'],
    master_products=master_products,
    month=7,
    year=2026,
    output_path=output_buffer
)
output_buffer.seek(0)
with open('temp_generated.xlsx', 'wb') as f:
    f.write(output_buffer.read())

# Run Tahap 1 logic
wb_sumber = openpyxl.load_workbook('temp_generated.xlsx', data_only=True)
ws_sumber = wb_sumber.worksheets[0]

def clean_key(text):
    return re.sub(r'[^A-Z0-9]', '', str(text or "").upper())

sales_map = {}
for r in range(2, ws_sumber.max_row + 1):
    date_val = ws_sumber.cell(row=r, column=1).value
    prod_val = ws_sumber.cell(row=r, column=16).value
    qty_val = ws_sumber.cell(row=r, column=17).value
    
    print(f"Row {r} -> Date: {date_val}, Prod: {prod_val}, Qty: {qty_val}")
    
    if isinstance(date_val, (datetime, date)): # Need to import date? No, datetime includes date
        day_str = f"{date_val.day:02d}"
    else:
        match = re.search(r'^(\d{1,2})[\/\-]', str(date_val).strip())
        if match:
            day_str = f"{int(match.group(1)):02d}"
        else:
            day_str = None
            
    print(f"  Parsed day_str: {day_str}")
    
    try:
        qty = float(qty_val)
    except:
        qty = 0
        
    if day_str and prod_val:
        key = (day_str, clean_key(prod_val))
        sales_map[key] = sales_map.get(key, 0) + qty

print("\nFinal sales_map:")
print(sales_map)
