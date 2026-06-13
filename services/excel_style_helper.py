from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

def copy_row_properties(ws, src_row, dst_row):
    """Menyalin seluruh aspek visual dan struktural dari src_row ke dst_row."""
    # Kloning tinggi baris
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        
    # Salin style per sel
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)

def reapply_merged_cells_safe(ws, src_row, dst_row):
    """Menduplikasi behavior merge cells horizontal pada baris baru."""
    merged_ranges = list(ws.merged_cells.ranges)
    for r in merged_ranges:
        if r.bounds[1] == src_row and r.bounds[3] == src_row:
            # Temukan boundaries kolom dari merge horizontal baris asal
            start_col, _, end_col, _ = r.bounds
            ws.merge_cells(start_row=dst_row, start_column=start_col, end_row=dst_row, end_column=end_col)