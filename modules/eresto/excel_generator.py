"""
Excel Generator Module for eResto Analysis using openpyxl.

Takes cleaned DataFrame, master product list, month, and year,
and writes them to an Excel file with the specified formatting:
- Raw columns in A, P, Q (cols B:O are hidden)
- Master product list starting in column S
- Calendar dates as headers starting in column T
- Automatically populated SUMIFS formulas inside the matrix area
- Elegant styling (emerald/dark-green headers, borders, zebra formatting)
"""
import calendar
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config

def generate_excel(df, date_col, product_col, qty_col, master_products, month, year, output_path):
    """
    Generate the final formatted Excel file.
    
    Args:
        df (pd.DataFrame): Cleaned DataFrame with original columns.
        date_col (str): The Order Date column name in df.
        product_col (str): The Product column name in df.
        qty_col (str): The Paid Qty column name in df.
        master_products (list): Ordered list of master product names.
        month (int): Selected month (1-12).
        year (int): Selected year.
        output_path (str): Filepath to save the resulting Excel workbook.
    """
    # 1. Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Analysis"
    
    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True
    
    # 2. Write raw data to Columns A through R only.
    # Column S onwards is reserved exclusively for the matrix (master product + dates).
    # Hard-cap writing at ERESTO_SUMMARY_PRODUCT_COL - 1 columns (col R = index 18, 1-based).
    RAW_COL_LIMIT = config.ERESTO_SUMMARY_PRODUCT_COL - 1  # = 18 (col R)
    original_cols = df.columns.tolist()[:RAW_COL_LIMIT]

    for col_idx, col_name in enumerate(original_cols):
        cell = ws.cell(row=1, column=col_idx + 1)
        cell.value = col_name

    row_num = 2
    for r in df.itertuples(index=False):
        for col_idx in range(RAW_COL_LIMIT):
            val = r[col_idx]
            cell = ws.cell(row=row_num, column=col_idx + 1)
            if col_idx == config.ERESTO_COL_ORDER_DATE:
                if hasattr(val, 'strftime'):
                    cell.value = val.strftime('%d/%m/%Y')
                else:
                    cell.value = str(val)
            else:
                cell.value = val
        row_num += 1
        
    # 3. Hide Columns B through O (indices 2 to 15, which is column index 1 to 14 in 0-based)
    for col_idx in range(config.ERESTO_HIDE_COLS_START, config.ERESTO_HIDE_COLS_END + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].hidden = True
        
    # 3b. Clear all existing data in Column S onwards (from row 1 to max_row)
    # This removes any leaked raw data from previous writes before we paint the matrix.
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=config.ERESTO_SUMMARY_PRODUCT_COL):
        for cell in row:
            cell.value = None

    # 4. Write Master Product list in Column S (starting at S2)
    # S1 is the header: "Master Product"
    product_header_cell = ws.cell(row=1, column=config.ERESTO_SUMMARY_PRODUCT_COL)
    product_header_cell.value = "Product"
    
    for idx, prod_name in enumerate(master_products):
        cell = ws.cell(row=idx + 2, column=config.ERESTO_SUMMARY_PRODUCT_COL)
        cell.value = prod_name
        
    # 5. Generate dynamic calendar dates starting in Column T (T1, U1, V1, etc.)
    # Based on selected month and year
    _, num_days = calendar.monthrange(year, month)
    
    date_cols_map = {} # day_number -> column_letter
    for day in range(1, num_days + 1):
        col_idx = config.ERESTO_DATE_START_COL + day - 1
        col_letter = get_column_letter(col_idx)
        
        # Write Date Header formatted as DD/MM/YYYY in Row 1
        cell = ws.cell(row=1, column=col_idx)
        current_date = date(year, month, day)
        cell.value = current_date.strftime('%d/%m/%Y')
        
        date_cols_map[day] = col_letter
        
    # 6. Generate Formula SUMIFS in each cell of the matrix
    # SUMIFS(sum_range, criteria_range1, criteria1, criteria_range2, criteria2)
    # - sum_range: Q:Q ($Q:$Q)
    # - criteria_range1: P:P ($P:$P)
    # - criteria1: S2 ($S2, $S3, etc. depending on product row)
    # - criteria_range2: A:A ($A:$A)
    # - criteria2: T$1 (T$1, U$1, V$1, etc. depending on column)
    
    # openpyxl uses 1-based index
    sum_range_letter = get_column_letter(config.ERESTO_COL_CLEANED_QTY + 1)   # Q
    criteria1_range_letter = get_column_letter(config.ERESTO_COL_PRODUCT + 1) # P
    criteria2_range_letter = get_column_letter(config.ERESTO_COL_ORDER_DATE + 1) # A
    product_col_letter = get_column_letter(config.ERESTO_SUMMARY_PRODUCT_COL) # S
    
    for row_idx, prod_name in enumerate(master_products, start=2):
        for day in range(1, num_days + 1):
            col_idx = config.ERESTO_DATE_START_COL + day - 1
            col_letter = get_column_letter(col_idx)
            
            # Formulate the cell value
            cell = ws.cell(row=row_idx, column=col_idx)
            
            formula = (
                f"=SUMIFS(${sum_range_letter}:${sum_range_letter},"
                f"${criteria1_range_letter}:${criteria1_range_letter},"
                f"${product_col_letter}{row_idx},"
                f"${criteria2_range_letter}:${criteria2_range_letter},"
                f"{col_letter}$1)"
            )
            cell.value = formula
            
    # 7. Apply Elegant Styling and Theme Colors
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_header = Font(name="Segoe UI", size=11, bold=True, color=config.COLOR_HEADER_FONT)
    font_date_header = Font(name="Segoe UI", size=10, bold=True, color=config.COLOR_DATE_HEADER_FONT)
    
    fill_header = PatternFill("solid", fgColor=config.COLOR_HEADER_BG)
    fill_date_header = PatternFill("solid", fgColor=config.COLOR_DATE_HEADER_BG)
    fill_product = PatternFill("solid", fgColor=config.COLOR_PRODUCT_BG)
    
    border_thin = Side(border_style="thin", color=config.COLOR_BORDER)
    border_thick = Side(border_style="medium", color=config.COLOR_HEADER_BG)
    
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Row 1 Headers (Raw Headers & Master Summary Headers)
    # For columns A, P, Q (and others)
    for col_idx in range(1, len(original_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    # Product Header (Column S)
    prod_header_cell = ws.cell(row=1, column=config.ERESTO_SUMMARY_PRODUCT_COL)
    prod_header_cell.font = font_header
    prod_header_cell.fill = fill_header
    prod_header_cell.alignment = align_left
    
    # Date Headers (Columns T onwards)
    for day in range(1, num_days + 1):
        col_idx = config.ERESTO_DATE_START_COL + day - 1
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_date_header
        cell.fill = fill_date_header
        cell.alignment = align_center
        # Set column width slightly wider for dates
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12
        
    # Set Width for product column S
    ws.column_dimensions[product_col_letter].width = 30
    
    # Style master product column and daily summary cells
    for row_idx, prod_name in enumerate(master_products, start=2):
        # Product cell S_row
        prod_cell = ws.cell(row=row_idx, column=config.ERESTO_SUMMARY_PRODUCT_COL)
        prod_cell.font = font_bold
        prod_cell.fill = fill_product
        prod_cell.alignment = align_left
        prod_cell.border = Border(left=border_thick, right=border_thin, top=border_thin, bottom=border_thin)
        
        # Zebra pattern logic for rows in matrix
        if row_idx % 2 == 0:
            fill_zebra = PatternFill("solid", fgColor=config.COLOR_ZEBRA_EVEN)
        else:
            fill_zebra = PatternFill("solid", fgColor=config.COLOR_ZEBRA_ODD)
            
        # Matrix cells (T_row to final_date_row)
        for day in range(1, num_days + 1):
            col_idx = config.ERESTO_DATE_START_COL + day - 1
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_normal
            cell.fill = fill_zebra
            cell.alignment = align_center
            cell.border = border_cell
            cell.number_format = '#,##0'
            
    # Set Raw Data columns widths (A, P, Q)
    ws.column_dimensions[get_column_letter(config.ERESTO_COL_ORDER_DATE + 1)].width = 15
    ws.column_dimensions[get_column_letter(config.ERESTO_COL_PRODUCT + 1)].width = 25
    ws.column_dimensions[get_column_letter(config.ERESTO_COL_CLEANED_QTY + 1)].width = 12
    
    # Apply a blank empty visual separator for Column R
    sep_letter = get_column_letter(config.ERESTO_SUMMARY_PRODUCT_COL - 1)
    ws.column_dimensions[sep_letter].width = 3
    
    # Freeze row 1 and column S (S = column 19, so freeze up to column R)
    ws.freeze_panes = "S2"
    
    # Save the file
    wb.save(output_path)
    return True