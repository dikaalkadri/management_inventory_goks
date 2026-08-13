import os
import re
import openpyxl
import pandas as pd
from datetime import datetime
import json
from openpyxl.styles import Font
from services.task_manager import active_tasks

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return re.sub(r'[^a-z0-9]', '', str(text).lower())

def process_perbaikan_data(task_id: str, so_bytes: bytes, inv_bytes: bytes, template_bytes: bytes, output_folder: str) -> dict:
    if task_id in active_tasks:
        active_tasks[task_id]['total'] = 100
        active_tasks[task_id]['progress'] = 10
        active_tasks[task_id]['current_item'] = "Membaca file yang diunggah..."
    
    # 1. Load Inventory Data for matching
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    inv_json_path = os.path.join(base_dir, 'data', 'inventory.json')
    inventory_items = []
    if os.path.exists(inv_json_path):
        try:
            with open(inv_json_path, 'r', encoding='utf-8') as f:
                inv_data_json = json.load(f)
                inventory_items = inv_data_json.get('items', [])
        except:
            pass

    # Load SO Drive alias mapping
    aliases_path = os.path.join(base_dir, 'data', 'perbaikan_data_aliases.json')
    so_drive_aliases = {}
    if os.path.exists(aliases_path):
        try:
            with open(aliases_path, 'r', encoding='utf-8') as f:
                so_drive_aliases = json.load(f).get('template_to_so', {})
        except:
            pass

    import tempfile
    import shutil
    
    tmpdir = tempfile.mkdtemp(prefix="perbaikan_data_")
    try:
        so_tmp = os.path.join(tmpdir, "so.xlsx")
        inv_tmp = os.path.join(tmpdir, "inv.xls")
        tpl_tmp = os.path.join(tmpdir, "template.xlsx")
        
        with open(so_tmp, 'wb') as f: f.write(so_bytes)
        with open(inv_tmp, 'wb') as f: f.write(inv_bytes)
        with open(tpl_tmp, 'wb') as f: f.write(template_bytes)
        
        if task_id in active_tasks:
            active_tasks[task_id]['progress'] = 30
            active_tasks[task_id]['current_item'] = "Mengekstrak data SO Drive & Inventory..."
            
        # Read SO Drive
        so_df = pd.DataFrame()
        try:
            so_df = pd.read_excel(so_tmp, sheet_name='REKAP', engine='openpyxl')
        except:
            try:
                so_df = pd.read_excel(so_tmp, sheet_name=0, engine='openpyxl') # Fallback
            except:
                pass
            
        so_data = []
        if not so_df.empty:
            for idx, row in so_df.iterrows():
                if len(row) >= 9:
                    nama = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                    
                    try: qty = float(row.iloc[8]) if pd.notna(row.iloc[8]) else 0.0
                    except: qty = 0.0
                    
                    try: awal = float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0.0
                    except: awal = 0.0
                    
                    try: masuk = float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0.0
                    except: masuk = 0.0

                    if nama and nama != "Nama Barang":
                        so_data.append({"nama": nama, "qty": qty, "awal": awal, "masuk": masuk})
                        
        # Read Inv Move
        inv_df = pd.DataFrame()
        try:
            inv_df = pd.read_excel(inv_tmp, engine='xlrd')
        except:
            try:
                inv_df = pd.read_excel(inv_tmp, engine='openpyxl')
            except:
                pass
                
        inv_data = {}
        if not inv_df.empty:
            sku_col = None
            name_col = None
            unit_col = None
            awal_col = None
            masuk_col = None
            
            for col in inv_df.columns:
                col_clean = str(col).lower()
                if "sku" in col_clean: sku_col = col
                elif "name" in col_clean and "inv" in col_clean: name_col = col
                elif "unit" in col_clean and "cost" not in col_clean: unit_col = col
                elif "opening" in col_clean or "awal" in col_clean: awal_col = col
                elif col_clean == "in" or " in " in f" {col_clean} " or "masuk" in col_clean: masuk_col = col
                
            if not sku_col: sku_col = inv_df.columns[0]
            if not name_col: name_col = inv_df.columns[1]
            if not unit_col: unit_col = inv_df.columns[8] if len(inv_df.columns) > 8 else None
            if not awal_col: awal_col = inv_df.columns[4] if len(inv_df.columns) > 4 else None
            if not masuk_col: masuk_col = inv_df.columns[5] if len(inv_df.columns) > 5 else None
            
            for idx, row in inv_df.iterrows():
                sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
                name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                unit = str(row[unit_col]).strip().lower() if unit_col and pd.notna(row[unit_col]) else ""
                
                try: awal = float(row[awal_col]) if awal_col and pd.notna(row[awal_col]) else 0.0
                except: awal = 0.0
                
                try: masuk = float(row[masuk_col]) if masuk_col and pd.notna(row[masuk_col]) else 0.0
                except: masuk = 0.0
                
                # Current Stock = Kolom H (index 7)
                current_stock_col = inv_df.columns[7] if len(inv_df.columns) > 7 else None
                try: current_stock = float(row[current_stock_col]) if current_stock_col and pd.notna(row[current_stock_col]) else 0.0
                except: current_stock = 0.0
                
                if sku and sku != "SKU":
                    inv_data[sku] = {"name": name, "unit": unit, "awal": awal, "masuk": masuk, "current_stock": current_stock, "row_data": row.tolist()}
                    
        if task_id in active_tasks:
            active_tasks[task_id]['progress'] = 60
            active_tasks[task_id]['current_item'] = "Memproses Template Adjustment..."
            
        # Process Template
        wb = openpyxl.load_workbook(tpl_tmp)
        ws = wb.active 
        
        header_row = 1
        sku_col_idx = 1
        name_col_idx = 2
        unit_col_idx = 3
        qty_col_idx = 5
        
        for r in range(1, 10):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, 10) if ws.cell(row=r, column=c).value]
            if "sku" in row_vals and "name" in row_vals:
                header_row = r
                for c in range(1, 15):
                    val = str(ws.cell(row=r, column=c).value).lower()
                    if "sku" in val: sku_col_idx = c
                    elif val == "name" or val == "inventory name": name_col_idx = c
                    elif "unit" in val: unit_col_idx = c
                    elif "qty" in val or "real" in val: qty_col_idx = c
                break
                
        # Set up headers for columns G, H, I, J, K, M, N
        ws.cell(row=header_row, column=7).value = "Stock Awal Drive"
        ws.cell(row=header_row, column=8).value = "Stock Masuk Drive"
        ws.cell(row=header_row, column=9).value = "Stock Opening Inventory Movement"
        ws.cell(row=header_row, column=10).value = "Stock In Inventory Movement"
        ws.cell(row=header_row, column=11).value = "Status"
        ws.cell(row=header_row, column=13).value = "Stock Akhir SO Drive"
        ws.cell(row=header_row, column=14).value = "Current Stock Inventory Movement"
                
        existing_skus = set()
        last_row = header_row
        
        def find_so_item(tpl_name):
            clean_tpl = clean_text(tpl_name)
            
            # 1. Exact match
            for item in so_data:
                if clean_text(item['nama']) == clean_tpl:
                    return item
            
            # 2. Alias mapping (SO Drive alias file)
            clean_tpl_spaced = re.sub(r'[^a-z0-9 ]', '', tpl_name.lower()).strip()
            alias_list = so_drive_aliases.get(clean_tpl_spaced, [])
            for alias in alias_list:
                alias_clean = clean_text(alias)
                for item in so_data:
                    if clean_text(item['nama']) == alias_clean:
                        return item
            
            # 3. Match via inventory.json keywords
            aliases_inv = [clean_tpl]
            for inv_item in inventory_items:
                if clean_text(inv_item.get('name')) == clean_tpl or clean_text(inv_item.get('display')) == clean_tpl:
                    for kw in inv_item.get('pdf_keywords', []):
                        aliases_inv.append(clean_text(kw))
            
            for item in so_data:
                cl_so = clean_text(item['nama'])
                if cl_so in aliases_inv:
                    return item
                    
            # 4. Substring match
            for item in so_data:
                cl_so = clean_text(item['nama'])
                if len(clean_tpl) > 3 and (clean_tpl in cl_so or cl_so in clean_tpl):
                    return item
                    
            return None
            
        def convert_qty(qty, so_name, tpl_unit):
            if qty is None:
                return ""
            
            tpl_unit_clean = clean_text(tpl_unit)
            if tpl_unit_clean == 'kg':
                if qty > 50: # Likely in grams
                    return qty / 1000.0
            elif tpl_unit_clean == 'liter' or tpl_unit_clean == 'l':
                if qty > 50: # Likely in ml
                    return qty / 1000.0
            
            return qty

        for r in range(header_row + 1, ws.max_row + 2):
            sku_val = ws.cell(row=r, column=sku_col_idx).value
            if sku_val:
                sku_str = str(sku_val).strip()
                existing_skus.add(sku_str)
                name_val = str(ws.cell(row=r, column=name_col_idx).value or "")
                unit_val = str(ws.cell(row=r, column=unit_col_idx).value or "")
                
                so_item = find_so_item(name_val)
                so_qty = so_item['qty'] if so_item else None
                so_nama = so_item['nama'] if so_item else None
                
                # Kolom E (Real Qty) dibiarkan kosong — tidak diisi otomatis
                
                # Fetch Inventory Movement details
                inv_info = inv_data.get(sku_str)
                inv_awal = inv_info['awal'] if inv_info else 0.0
                inv_masuk = inv_info['masuk'] if inv_info else 0.0
                inv_current = inv_info['current_stock'] if inv_info else 0.0
                
                so_awal = so_item['awal'] if so_item else 0.0
                so_masuk = so_item['masuk'] if so_item else 0.0
                
                # Write to G, H, I, J
                ws.cell(row=r, column=7).value = so_awal
                ws.cell(row=r, column=8).value = so_masuk
                ws.cell(row=r, column=9).value = inv_awal
                ws.cell(row=r, column=10).value = inv_masuk
                
                # Write Status to K
                if round(so_awal, 4) == round(inv_awal, 4) and round(so_masuk, 4) == round(inv_masuk, 4):
                    ws.cell(row=r, column=11).value = "✅"
                    ws.cell(row=r, column=11).font = Font(name='Segoe UI Emoji', color='00B050')
                else:
                    ws.cell(row=r, column=11).value = "❌"
                    ws.cell(row=r, column=11).font = Font(name='Segoe UI Emoji', color='FF0000')
                
                # Write to M (Stock Akhir SO Drive) dan N (Current Stock Inv Move)
                so_akhir = convert_qty(so_qty, so_nama, unit_val)
                ws.cell(row=r, column=13).value = so_akhir if so_akhir != "" else None
                ws.cell(row=r, column=14).value = inv_current
                    
                last_row = r
                
        if task_id in active_tasks:
            active_tasks[task_id]['progress'] = 80
            active_tasks[task_id]['current_item'] = "Menambahkan item baru dari Inventory..."
            
        for inv_sku, inv_info in inv_data.items():
            if inv_sku not in existing_skus:
                last_row += 1
                name_val = inv_info['name']
                unit_val = inv_info['unit']
                
                ws.cell(row=last_row, column=sku_col_idx).value = inv_sku
                ws.cell(row=last_row, column=name_col_idx).value = name_val
                ws.cell(row=last_row, column=unit_col_idx).value = unit_val
                
                if len(inv_info['row_data']) > 2:
                    ws.cell(row=last_row, column=4).value = inv_info['row_data'][2]
                
                so_item = find_so_item(name_val)
                so_qty = so_item['qty'] if so_item else None
                so_nama = so_item['nama'] if so_item else None
                
                # Kolom E (Real Qty) dibiarkan kosong
                
                inv_awal = inv_info['awal']
                inv_masuk = inv_info['masuk']
                inv_current = inv_info['current_stock']
                so_awal = so_item['awal'] if so_item else 0.0
                so_masuk = so_item['masuk'] if so_item else 0.0
                
                ws.cell(row=last_row, column=7).value = so_awal
                ws.cell(row=last_row, column=8).value = so_masuk
                ws.cell(row=last_row, column=9).value = inv_awal
                ws.cell(row=last_row, column=10).value = inv_masuk
                
                if round(so_awal, 4) == round(inv_awal, 4) and round(so_masuk, 4) == round(inv_masuk, 4):
                    ws.cell(row=last_row, column=11).value = "✅"
                    ws.cell(row=last_row, column=11).font = Font(name='Segoe UI Emoji', color='00B050')
                else:
                    ws.cell(row=last_row, column=11).value = "❌"
                    ws.cell(row=last_row, column=11).font = Font(name='Segoe UI Emoji', color='FF0000')
                
                # Write to M dan N
                so_akhir = convert_qty(so_qty, so_nama, unit_val)
                ws.cell(row=last_row, column=13).value = so_akhir if so_akhir != "" else None
                ws.cell(row=last_row, column=14).value = inv_current

        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_filename = f"Template_Adjustment_Terisi_{now_str}.xlsx"
        out_filepath = os.path.join(output_folder, out_filename)
        wb.save(out_filepath)
        
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    if task_id in active_tasks:
        active_tasks[task_id]['progress'] = 100
        active_tasks[task_id]['current_item'] = "Selesai!"

    return {
        "filename": out_filename,
        "download_url": f"/api/perbaikan-data/download/{out_filename}"
    }
