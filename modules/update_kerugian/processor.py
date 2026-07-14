import os
import re
import zipfile
import tempfile
import shutil
from io import BytesIO
from datetime import datetime, date
import openpyxl

from modules.sinkronisasi.processor import proses_sinkronisasi_excel
from modules.mass_sinko.processor import _build_outlet_map, _to_xlsx

def _get_available_days(eresto_xlsx):
    wb = openpyxl.load_workbook(eresto_xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    days = set()
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r < 2:
            continue
        if not row:
            continue
        date_val = row[0]
        if date_val is None:
            continue
        day_str = None
        if isinstance(date_val, (datetime, date)):
            day_str = f"{date_val.day:02d}"
        else:
            match = re.search(r'^(\d{1,2})[\/\-]', str(date_val).strip())
            if match:
                day_str = f"{int(match.group(1)):02d}"
        if day_str:
            days.add(day_str)
    wb.close()
    return days

def process_update_kerugian(eresto_zip_bytes: bytes, so_zip_bytes: bytes, kerugian_zip_bytes: bytes, output_folder: str) -> dict:
    # 1. Ekstrak ketiga ZIP
    eresto_map = _build_outlet_map(eresto_zip_bytes, "eResto")
    so_map     = _build_outlet_map(so_zip_bytes, "SO")
    kerugian_map = _build_outlet_map(kerugian_zip_bytes, "Kerugian")

    # Kita hanya memproses outlet yang ada di Kerugian (Bahan 3)
    all_nos = sorted(kerugian_map.keys())

    missing_so = []
    missing_eresto = []
    errors = []
    success_outlets = []

    tmpdir = tempfile.mkdtemp(prefix="update_kerugian_")

    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _proses_single(no):
            kerugian_info = kerugian_map[no]
            outlet_name = kerugian_info['filename'].split('_')[1] if len(kerugian_info['filename'].split('_')) > 1 else str(no)

            if no not in eresto_map:
                return {"status": "missing_eresto", "no": no}

            if no not in so_map:
                return {"status": "missing_so", "no": no}

            eresto_info = eresto_map[no]
            so_info = so_map[no]

            try:
                # 2. Simpan file sementara
                eresto_tmp = os.path.join(tmpdir, f"eresto_{no}_{eresto_info['filename']}")
                so_tmp = os.path.join(tmpdir, f"so_{no}_{so_info['filename']}")
                kerugian_tmp = os.path.join(tmpdir, f"kerugian_{no}_{kerugian_info['filename']}")

                with open(eresto_tmp, 'wb') as f:
                    f.write(eresto_info['bytes'])
                with open(so_tmp, 'wb') as f:
                    f.write(so_info['bytes'])
                with open(kerugian_tmp, 'wb') as f:
                    f.write(kerugian_info['bytes'])

                eresto_xlsx = _to_xlsx(eresto_tmp, tmpdir)
                so_xlsx = _to_xlsx(so_tmp, tmpdir)
                kerugian_xlsx = _to_xlsx(kerugian_tmp, tmpdir)

                # 3. Ambil daftar tanggal dari eResto
                available_days = _get_available_days(eresto_xlsx)

                # 4. Patch (salin E,F,G,H) dari SO ke Kerugian
                wb_tujuan = openpyxl.load_workbook(kerugian_xlsx)
                wb_so = openpyxl.load_workbook(so_xlsx, read_only=True, data_only=True)

                patched = False
                for day_str in available_days:
                    if day_str in wb_tujuan.sheetnames and day_str in wb_so.sheetnames:
                        ws_t = wb_tujuan[day_str]
                        ws_s = wb_so[day_str]
                        # Salin kolom E(5), F(6), G(7), H(8) untuk baris 2 hingga 120 secara cepat menggunakan iter_rows
                        for r_idx, row_cells in enumerate(ws_s.iter_rows(min_row=2, max_row=120, min_col=5, max_col=8, values_only=True), start=2):
                            for c_idx, val in enumerate(row_cells, start=5):
                                if val is not None:
                                    ws_t.cell(row=r_idx, column=c_idx).value = val
                        patched = True

                wb_so.close()

                if patched:
                    wb_tujuan.save(kerugian_xlsx)
                else:
                    wb_tujuan.close()

                # 5. Lempar ke prosesor utama (menggunakan Kerugian yang sudah di-patch sebagai tujuan)
                output_filename, sheets_count = proses_sinkronisasi_excel(
                    sumber_path_xlsx=eresto_xlsx,
                    tujuan_path_xlsx=kerugian_xlsx,
                    output_folder=output_folder,
                )

                return {
                    "status": "success",
                    "no": no,
                    "name": outlet_name,
                    "output_file": output_filename,
                    "sheets": sheets_count
                }

            except Exception as e:
                return {
                    "status": "error",
                    "name": outlet_name,
                    "error": str(e)
                }

        # Jalankan secara paralel menggunakan ThreadPoolExecutor dengan 5 worker
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(_proses_single, no): no for no in all_nos}
            for future in as_completed(futures):
                res = future.result()
                status = res["status"]
                if status == "success":
                    success_outlets.append({
                        "no": res["no"],
                        "name": res["name"],
                        "output_file": res["output_file"],
                        "sheets": res["sheets"]
                    })
                    print(f"[UPDATE-KERUGIAN] ✅ {res['name']} → {res['output_file']} ({res['sheets']} sheets patched)")
                elif status == "missing_eresto":
                    missing_eresto.append(f"{res['no']} (tidak ada di eResto)")
                elif status == "missing_so":
                    missing_so.append(f"{res['no']} (tidak ada di SO)")
                elif status == "error":
                    errors.append({"outlet": res["name"], "error": res["error"]})
                    print(f"[UPDATE-KERUGIAN] ❌ {res['name']}: {res['error']}")

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    # 6. Buat ZIP Output
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"Update_Kerugian_{now_str}.zip"
    zip_path = os.path.join(output_folder, zip_name)

    processed_count = len(success_outlets)
    if processed_count > 0:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in success_outlets:
                src = os.path.join(output_folder, item['output_file'])
                if os.path.exists(src):
                    zout.write(src, arcname=item['output_file'])

    return {
        "zip_filename": zip_name if processed_count > 0 else None,
        "processed": processed_count,
        "missing_so": missing_so,
        "missing_eresto": missing_eresto,
        "errors": errors,
        "success_outlets": [f"{i['no']}. {i['name']}" for i in success_outlets],
    }
