import os
import re
import datetime
import zipfile
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.cell.cell import MergedCell

from modules.stockin.helpers import convert_to_xlsx_if_needed
from services.task_manager import active_tasks

# Border tipis hanya untuk kolom A s/d M pada baris bahan yang ditambahkan
BORDER_THIN = Side(border_style="thin", color="000000")
CELL_BORDER_AM = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)

def _safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    try:
        cell.value = value
    except:
        pass
    return cell

def proses_mass_update(task_id, file_paths, formulas, output_folder):
    """
    Memproses daftar file Excel:
    - Mengubah rumus di Kolom J sesuai data 'formulas'
    - Menulis kategori di Kolom A, nama bahan di Kolom B, dan harga di Kolom W
    - Menambahkan border grid HANYA pada Kolom A s/d M (kolom 1-13) untuk baris yang di-update
    - Kolom N sampai X tidak diganggu / tidak dirubah
    - Mengemas hasilnya ke dalam file ZIP
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder, exist_ok=True)
        
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"MassUpdate_Rumus_{timestamp}.zip"
    zip_path = os.path.join(output_folder, zip_filename)
    
    success_count = 0
    fail_count = 0
    errors = []
    
    if task_id in active_tasks:
        active_tasks[task_id]['total'] = len(file_paths)
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for fpath in file_paths:
            basename = os.path.basename(fpath)
            clean_basename = basename[3:] if basename.startswith("mu_") else basename
            
            if task_id in active_tasks:
                active_tasks[task_id]['current_item'] = f"Memproses: {clean_basename}"
            
            try:
                fpath_xlsx = convert_to_xlsx_if_needed(fpath)
                if not fpath_xlsx or not os.path.exists(fpath_xlsx):
                    raise Exception("Gagal memproses file atau format tidak didukung")
                
                wb = openpyxl.load_workbook(fpath_xlsx)
                sheet_updated = False
                
                from services.catalog_service import load_materials
                mat_list = load_materials()
                mat_dict = {m.get('name', '').strip().lower(): m.get('price', 0) for m in mat_list}

                for sheet_name in wb.sheetnames:
                    # Proses sheet harian (01, 02, ..., 31)
                    if re.match(r"^\d{2}$", sheet_name.strip()):
                        ws = wb[sheet_name]
                        ws.protection.sheet = False

                        for row_str, item in formulas.items():
                            try:
                                row_idx = int(row_str)
                                formula_val = item.get("formula", "")
                                orig_cat = item.get("category", "").strip()
                                orig_name = item.get("name", "").strip()
                                mat_name = orig_name.lower()
                                
                                # 1. Set border HANYA untuk kolom A s/d M (kolom 1 s/d 13) pada baris ini
                                for col_idx in range(1, 14):
                                    c = ws.cell(row=row_idx, column=col_idx)
                                    if not isinstance(c, MergedCell):
                                        c.border = CELL_BORDER_AM
                                
                                # 2. Tulis Kategori (Kolom A)
                                if orig_cat:
                                    _safe_write(ws, row_idx, 1, orig_cat)
                                    
                                # 3. Tulis Nama Bahan (Kolom B)
                                if orig_name:
                                    _safe_write(ws, row_idx, 2, orig_name)
                                    
                                # 4. Tulis Rumus J (Kolom J)
                                if formula_val:
                                    _safe_write(ws, row_idx, 10, formula_val)
                                    
                                # 5. Tulis Master Harga (Kolom W)
                                if mat_name and mat_name in mat_dict:
                                    price_val = mat_dict[mat_name]
                                    _safe_write(ws, row_idx, 23, price_val)
                                elif item.get("price") is not None:
                                    _safe_write(ws, row_idx, 23, item.get("price"))
                            except ValueError:
                                continue

                        # Kembalikan proteksi sheet
                        ws.protection.sheet = True
                        ws.protection.password = "12345678"
                        sheet_updated = True
                
                if sheet_updated:
                    temp_output_path = os.path.join(output_folder, f"temp_{clean_basename}")
                    wb.save(temp_output_path)
                    wb.close()
                    
                    zipf.write(temp_output_path, clean_basename)
                    try:
                        os.remove(temp_output_path)
                    except:
                        pass
                        
                    success_count += 1
                else:
                    wb.close()
                    fail_count += 1
                    errors.append(f"{clean_basename}: Tidak ada sheet harian (01-31) yang ditemukan.")
                    
                if fpath_xlsx != fpath and os.path.exists(fpath_xlsx):
                    try:
                        os.remove(fpath_xlsx)
                    except:
                        pass
                        
            except Exception as e:
                fail_count += 1
                errors.append(f"{clean_basename}: {str(e)}")

            if task_id in active_tasks:
                active_tasks[task_id]['progress'] += 1
                
        # Bersihkan file origin
        for p in file_paths:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass
                
    return zip_filename, success_count, fail_count, errors
