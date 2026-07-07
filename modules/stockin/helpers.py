"""
Helper functions for Stock-In Manager.
Handles inventory loading/saving, app config loading/saving, excel path determination,
workbook retrieval, sheet setup, item section management, duplicate checking, sisa stok calculation,
recalculation, and saving.
"""
import os
import json
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from xls2xlsx import XLS2XLSX
import openpyxl
from datetime import datetime, timedelta

def load_inventory():
    with open('inventory.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def save_inventory(data):
    with open('inventory.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_app_config():
    cfg_path = 'app_config.json'
    if not os.path.exists(cfg_path):
        return {"excel_filename": "", "excel_folder": "", "stok_awal": {}}
    with open(cfg_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_app_config(data):
    with open('app_config.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def convert_to_xlsx_if_needed(filepath):
    if filepath.lower().endswith('.xls'):
        try:
            x2x = XLS2XLSX(filepath)
            new_filepath = filepath + "x"
            x2x.to_xlsx(new_filepath)
            return new_filepath
        except Exception as e:
            raise ValueError(f"Gagal mengonversi file .xls: {str(e)}")
    return filepath

def safe_save_workbook(wb, excel_path):
    folder = os.path.dirname(excel_path) or '.'
    basename = os.path.basename(excel_path)
    name, ext = os.path.splitext(basename)
    tmp_path = os.path.join(folder, f"{name}_tmp{ext}")
    try:
        wb.save(tmp_path)
    except Exception as e:
        return False, f"Gagal menulis file sementara: {e}"
    try:
        if os.path.exists(excel_path):
            os.replace(tmp_path, excel_path)
        else:
            os.rename(tmp_path, excel_path)
        return True, "ok"
    except PermissionError:
        try:
            os.remove(tmp_path)
        except:
            pass
        return False, "⚠️ File Excel sedang terbuka! Tutup Excel terlebih dahulu, lalu simpan ulang."
    except Exception as e:
        try:
            os.remove(tmp_path)
        except:
            pass
        return False, f"Gagal menyimpan: {e}"

def get_excel_path():
    cfg = load_app_config()
    folder = cfg.get('excel_folder', 'data')
    filename = cfg.get('excel_filename', '')
    if not filename:
        return None, "File Excel belum dikonfigurasi"
    return os.path.join(folder, filename), None

def get_or_create_workbook(path):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
            
        inv = load_inventory()
        cfg = load_app_config()
        filename = cfg.get('excel_filename', '')
        stok_awal_map = cfg.get('stok_awal', {}).get(filename, {})
        
        for item_cfg in inv.get('items', []):
            item_name = item_cfg.get('display', item_cfg.get('name', ''))
            sheet_name = item_cfg.get('sheet', 'Bahan Lainnya')
            ws = get_or_create_sheet(wb, sheet_name, item_name)
            _, _, _, start_col = find_or_create_item_section(ws, item_name)
            
            stok_awal = float(stok_awal_map.get(item_name, 0))
            _write_stock_awal_row(ws, start_col, 5, stok_awal)
            recalculate_item_stock(ws, start_col, 5)
            
    return wb

def get_or_create_sheet(wb, sheet_name, item_name):
    target = sheet_name
    if target not in wb.sheetnames:
        ws = wb.create_sheet(target)
        ws.append(["STOCK OPNAME - " + target])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([])
    return wb[target]

def find_or_create_item_section(ws, item_name):
    headers = ["Tanggal", "Masuk", "Keluar", "Sisa Stok", "Keterangan"]
    search_limit = max(500, ws.max_column + 10)
    for start_col in range(1, search_limit, 6):
        cell = ws.cell(row=3, column=start_col)
        if cell.value and str(cell.value).strip().upper() == item_name.upper():
            col_header_row = 4
            last_data_row = col_header_row
            last_sisa = 0
            for r in range(5, ws.max_row + 1):
                if ws.cell(row=r, column=start_col).value is None:
                    break
                last_data_row = r
                sisa_cell = ws.cell(row=r, column=start_col + 3)
                if sisa_cell.value is not None:
                    try:
                        last_sisa = float(sisa_cell.value)
                    except:
                        pass
            return col_header_row, last_data_row, last_sisa, start_col

    start_col = 1
    while ws.cell(row=3, column=start_col).value is not None:
        start_col += 6

    ws.cell(row=3, column=start_col, value=item_name)
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 4)
    header_cell = ws.cell(row=3, column=start_col)
    header_cell.font = Font(bold=True, size=11, color="FFFFFF")
    header_cell.fill = PatternFill("solid", fgColor="2D6A4F")
    header_cell.alignment = Alignment(horizontal='center')

    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=start_col + i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D8F3DC")
        cell.alignment = Alignment(horizontal='center')

    widths = [14, 10, 10, 12, 30]
    for i, width in enumerate(widths):
        col_letter = get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = width

    return 4, 4, 0, start_col

def check_duplicate(ws, col_header_row, tanggal_str, masuk_keluar, keterangan, start_col):
    for r in range(col_header_row + 1, ws.max_row + 1):
        cell_tgl = ws.cell(row=r, column=start_col)
        if cell_tgl.value is None:
            break
        if parse_date_to_ymd(cell_tgl.value) != parse_date_to_ymd(tanggal_str):
            continue
        if masuk_keluar == "Masuk" and ws.cell(row=r, column=start_col + 1).value in (None, ""):
            continue
        if masuk_keluar == "Keluar" and ws.cell(row=r, column=start_col + 2).value in (None, ""):
            continue
        ket_cell = str(ws.cell(row=r, column=start_col + 4).value or "").strip()
        if ket_cell == str(keterangan or "").strip():
            return r
    return None

def _write_stock_awal_row(ws, start_col, stock_row, stok_awal):
    for c in range(start_col, start_col + 5):
        cell = ws.cell(row=stock_row, column=c)
        cell.fill = PatternFill("solid", fgColor="9D130F")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=stock_row, column=start_col).value = ""
    ws.cell(row=stock_row, column=start_col + 1).value = None
    ws.cell(row=stock_row, column=start_col + 2).value = None
    ws.cell(row=stock_row, column=start_col + 3).value = float(stok_awal)
    ws.cell(row=stock_row, column=start_col + 4).value = "STOCK GUDANG AWAL"

def parse_date_to_ymd(date_val):
    if not date_val: return ""
    date_str = str(date_val).strip()
    try:
        if len(date_str) >= 10:
            if date_str[2] == '-' and date_str[5] == '-':
                return datetime.strptime(date_str[:10], "%d-%m-%Y").strftime("%Y-%m-%d")
            elif date_str[4] == '-' and date_str[7] == '-':
                return date_str[:10]
    except:
        pass
    return date_str

def format_date_to_dmy(date_val):
    if not date_val: return ""
    date_str = str(date_val).strip()
    try:
        if len(date_str) >= 10 and date_str[4] == '-' and date_str[7] == '-':
            return datetime.strptime(date_str[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        pass
    return date_str

def recalculate_item_stock(ws, start_col, stock_row=5):
    # 1. Kumpulkan semua data transaksi dari baris 5 ke bawah
    transactions = []
    for r in range(stock_row, ws.max_row + 2):
        tanggal = parse_date_to_ymd(ws.cell(row=r, column=start_col).value)
        masuk   = ws.cell(row=r, column=start_col + 1).value
        keluar  = ws.cell(row=r, column=start_col + 2).value
        sisa    = ws.cell(row=r, column=start_col + 3).value
        ket     = ws.cell(row=r, column=start_col + 4).value

        # Abaikan baris kosong
        if all(v in (None, "") for v in [tanggal, masuk, keluar, sisa, ket]):
            continue
            
        transactions.append({
            'tanggal': tanggal,
            'masuk': masuk,
            'keluar': keluar,
            'sisa_asli': sisa,  # hanya dipakai untuk stock awal
            'ket': ket
        })
        
        # Kosongkan baris agar bersih sebelum ditulis ulang
        for c in range(start_col, start_col + 5):
            ws.cell(row=r, column=c).value = None

    # Pisahkan data stock awal dan transaksi
    data_transaksi = [t for t in transactions if str(t['ket'] or "").strip().upper() != "STOCK GUDANG AWAL"]
    stock_awal_trans = [t for t in transactions if str(t['ket'] or "").strip().upper() == "STOCK GUDANG AWAL"]

    # Kalkulasi H-1 Bulan untuk Stock Awal
    target_date = ""
    if data_transaksi:
        valid_dates = [t['tanggal'] for t in data_transaksi if t['tanggal'] and str(t['tanggal']).strip()]
        if valid_dates:
            min_date_str = min(valid_dates)
            try:
                min_date = datetime.strptime(str(min_date_str), "%Y-%m-%d")
                first_day_of_month = min_date.replace(day=1)
                last_day_prev_month = first_day_of_month - timedelta(days=1)
                target_date = last_day_prev_month.strftime("%Y-%m-%d")
            except Exception:
                pass # Tetap string kosong jika gagal parse

    for t in stock_awal_trans:
        t['tanggal'] = target_date

    # Gabungkan kembali
    transactions = stock_awal_trans + data_transaksi

    # 2. Sortir berdasarkan tanggal (kosong "" akan berada paling atas, lalu tanggal terlama ke terbaru)
    transactions.sort(key=lambda x: str(x['tanggal'] or ""))

    # 3. Tulis ulang ke Excel dengan format yang benar dan Rumus Sisa
    current_row = stock_row
    for t in transactions:
        is_stock_awal = str(t['ket'] or "").strip().upper() == "STOCK GUDANG AWAL"
        
        ws.cell(row=current_row, column=start_col).value = format_date_to_dmy(t['tanggal'])
        ws.cell(row=current_row, column=start_col + 1).value = t['masuk']
        ws.cell(row=current_row, column=start_col + 2).value = t['keluar']
        ws.cell(row=current_row, column=start_col + 4).value = t['ket']

        # Terapkan warna
        fill_color = "FFFFFF"
        font_color = "000000"
        is_bold = False

        if is_stock_awal:
            fill_color = "9D130F"  # Merah
            font_color = "FFFFFF"
            is_bold = True
        elif t['masuk'] not in (None, "", 0, "0"):
            fill_color = "FFEA00"  # Kuning
        else:
            fill_color = "F0FFF4" if current_row % 2 == 0 else "FFFFFF"

        for c in range(start_col, start_col + 5):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(bold=is_bold, color=font_color)
            cell.alignment = Alignment(horizontal="center")

        # Injeksi Rumus Excel untuk Sisa Stock
        if is_stock_awal or current_row == stock_row:
            # Baris pertama (Stock Awal) pakai angka absolut
            ws.cell(row=current_row, column=start_col + 3).value = float(t['sisa_asli'] or 0)
        else:
            sisa_col = get_column_letter(start_col + 3)
            masuk_col = get_column_letter(start_col + 1)
            keluar_col = get_column_letter(start_col + 2)
            
            # Rumus: Sisa Atas + Masuk - Keluar
            prev_sisa_cell = f"{sisa_col}{current_row - 1}"
            masuk_cell = f"{masuk_col}{current_row}"
            keluar_cell = f"{keluar_col}{current_row}"
            
            formula = f"={prev_sisa_cell}+IF(ISNUMBER({masuk_cell}),{masuk_cell},0)-IF(ISNUMBER({keluar_cell}),{keluar_cell},0)"
            ws.cell(row=current_row, column=start_col + 3).value = formula

        current_row += 1
