"""
Flask Blueprint Routes for Stock-In Manager.
Handles page rendering and API endpoints for stock operations.
"""
from flask import Blueprint, request, jsonify, render_template
import os
import re
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Alignment

from modules.stockin.helpers import (
    load_inventory,
    save_inventory,
    load_app_config,
    save_app_config,
    get_excel_path,
    get_or_create_workbook,
    get_or_create_sheet,
    find_or_create_item_section,
    _write_stock_awal_row,
    recalculate_item_stock,
    check_duplicate,
    safe_save_workbook
)

stockin_bp = Blueprint('stockin', __name__, template_folder='../../templates')

@stockin_bp.route('/')
def index():
    return render_template('index.html')

@stockin_bp.route('/settings')
def settings():
    return render_template('settings.html')

@stockin_bp.route('/stock-gudang')
def stock_gudang():
    return render_template('stock_gudang.html')

@stockin_bp.route('/api/config', methods=['GET'])
def get_config():
    cfg = load_app_config()
    inv = load_inventory()
    return jsonify({**cfg, "outlets": inv.get("outlets", [])})

@stockin_bp.route('/api/inventory', methods=['GET'])
def get_inventory():
    inv = load_inventory()
    return jsonify({
        "items": inv.get("items", []),
        "outlets": inv.get("outlets", [])
    })

@stockin_bp.route('/api/inventory/item', methods=['POST'])
def add_inventory_item():
    data = request.json
    inv = load_inventory()
    if "items" not in inv:
        inv["items"] = []
    
    # Generate ID from name
    base_id = re.sub(r'[^a-z0-9]+', '_', data.get('name', '').lower()).strip('_')
    if not base_id:
        import uuid
        base_id = str(uuid.uuid4())[:8]
    
    item_id = base_id
    counter = 1
    # Ensure uniqueness
    while any(item.get('id') == item_id for item in inv["items"]):
        item_id = f"{base_id}_{counter}"
        counter += 1
        
    data['id'] = item_id
    inv["items"].append(data)
    save_inventory(inv)
    return jsonify({"status": "ok", "id": item_id})

@stockin_bp.route('/api/inventory/item/<item_id>', methods=['PUT'])
def update_inventory_item(item_id):
    data = request.json
    inv = load_inventory()
    for item in inv.get("items", []):
        if item.get("id") == item_id:
            item.update(data)
            # Ensure id doesn't get overridden by mistake
            item['id'] = item_id
            break
    save_inventory(inv)
    return jsonify({"status": "ok"})

@stockin_bp.route('/api/inventory/item/<item_id>', methods=['DELETE'])
def delete_inventory_item(item_id):
    inv = load_inventory()
    if "items" in inv:
        inv["items"] = [item for item in inv["items"] if item.get("id") != item_id]
        save_inventory(inv)
    return jsonify({"status": "ok"})

@stockin_bp.route('/api/inventory/outlets', methods=['PUT'])
def update_inventory_outlets():
    data = request.json
    inv = load_inventory()
    inv["outlets"] = data.get("outlets", [])
    save_inventory(inv)
    return jsonify({"status": "ok"})

@stockin_bp.route('/api/config', methods=['POST'])
def save_config():
    data = request.json
    cfg = load_app_config()
    if 'excel_filename' in data:
        cfg['excel_filename'] = data['excel_filename']
    if 'excel_folder' in data:
        cfg['excel_folder'] = data['excel_folder']
    save_app_config(cfg)
    return jsonify({"status": "ok"})

@stockin_bp.route('/api/config/check-file', methods=['POST'])
def check_file():
    data = request.json
    filename = data.get('filename', '').strip()
    folder = data.get('folder', '').strip()
    if not filename or not folder:
        return jsonify({"found": False, "message": "Nama file dan folder harus diisi"})
    full_path = os.path.join(folder, filename)
    found = os.path.isfile(full_path)
    return jsonify({
        "found": found,
        "path": full_path,
        "message": f"✅ Ditemukan: {full_path}" if found else f"❌ Tidak ditemukan: {full_path}"
    })

@stockin_bp.route('/api/config/stok-awal', methods=['GET'])
def get_stok_awal():
    cfg = load_app_config()
    filename = cfg.get('excel_filename', '')
    stok_awal_all = cfg.get('stok_awal', {})
    return jsonify({"filename": filename, "stok_awal": stok_awal_all.get(filename, {})})

@stockin_bp.route('/api/config/stok-awal', methods=['POST'])
def save_stok_awal():
    data = request.json
    cfg = load_app_config()
    filename = cfg.get('excel_filename', '')
    if not filename:
        return jsonify({"status": "error", "message": "File Excel belum dikonfigurasi"})
    if 'stok_awal' not in cfg:
        cfg['stok_awal'] = {}
    cfg['stok_awal'][filename] = data.get('stok_awal', {})
    save_app_config(cfg)

    excel_path, err = get_excel_path()
    if not err and os.path.exists(excel_path):
        try:
            wb = get_or_create_workbook(excel_path)
            stok_awal_map = cfg['stok_awal'][filename]
            inv = load_inventory()

            for item_cfg in inv['items']:
                item_name = item_cfg['display']
                sheet_name = item_cfg.get('sheet', 'Bahan Lainnya')
                ws = get_or_create_sheet(wb, sheet_name, item_name)
                col_header_row, last_data_row, last_sisa, start_col = find_or_create_item_section(ws, item_name)
                stock_row = 5
                stok_awal = float(stok_awal_map.get(item_name, 0))

                _write_stock_awal_row(ws, start_col, stock_row, stok_awal)
                recalculate_item_stock(ws, start_col, stock_row)

            ok, msg = safe_save_workbook(wb, excel_path)
            if not ok:
                return jsonify({"status": "error", "message": msg})
        except Exception as e:
            return jsonify({"status": "ok", "warning": f"Config tersimpan, tapi recalculate gagal: {str(e)}"})
    return jsonify({"status": "ok", "message": "✅ Stok awal disimpan dan histori Excel diperbarui"})

@stockin_bp.route('/api/recalculate-stock', methods=['POST'])
def recalculate_stock():
    excel_path, err = get_excel_path()
    if err:
        return jsonify({"status": "error", "message": err})
    if not os.path.exists(excel_path):
        return jsonify({"status": "error", "message": "Excel tidak ditemukan"})

    wb = get_or_create_workbook(excel_path)
    cfg = load_app_config()
    filename = cfg.get('excel_filename', '')
    stok_awal_map = cfg.get('stok_awal', {}).get(filename, {})
    inv = load_inventory()

    for item_cfg in inv['items']:
        item_name = item_cfg['display']
        sheet_name = item_cfg.get('sheet', 'Bahan Lainnya')
        ws = get_or_create_sheet(wb, sheet_name, item_name)
        col_header_row, last_data_row, last_sisa, start_col = find_or_create_item_section(ws, item_name)
        stock_row = 5
        stok_awal = float(stok_awal_map.get(item_name, 0))

        _write_stock_awal_row(ws, start_col, stock_row, stok_awal)
        recalculate_item_stock(ws, start_col, stock_row)

    ok, msg = safe_save_workbook(wb, excel_path)
    if not ok:
        return jsonify({"status": "error", "message": msg})
    return jsonify({"status": "ok", "message": "✅ Semua stock berhasil direcalculate"})

@stockin_bp.route('/api/submit', methods=['POST'])
def submit_to_excel():
    data = request.json
    cards = data.get('cards', [])
    excel_path, err = get_excel_path()
    if err:
        return jsonify({"status": "error", "message": err})
    
    os.makedirs(os.path.dirname(excel_path) if os.path.dirname(excel_path) else '.', exist_ok=True)
    wb = get_or_create_workbook(excel_path)
    
    written, errors = [], []
    for card in cards:
        outlet, tanggal, items = card.get('outlet', ''), card.get('tanggal', ''), card.get('items', [])
        for item in items:
            try:
                item_name = item.get('name', '')
                
                # --- FRUCTOSE MAPPING LOGIC ---
                item_name_lower = item_name.lower()
                if "fructose" in item_name_lower:
                    if "15000" in item_name_lower or "10000" in item_name_lower or "5000" in item_name_lower:
                        item_name = "Fructose (5000)"
                    elif "2000" in item_name_lower:
                        item_name = "Fructose (2000)"
                    elif "1000" in item_name_lower:
                        item_name = "Fructose (1000)"
                # ------------------------------
                
                sheet_name = item.get('sheet', 'Bahan Lainnya')
                masuk_keluar = item.get('masuk_keluar', 'Keluar')
                qty = item.get('qty_excel', 0)
                keterangan = item.get('keterangan', '') or f"{outlet}" if outlet else ""
                
                ws = get_or_create_sheet(wb, sheet_name, item_name)
                col_header_row, last_data_row, last_sisa, start_col = find_or_create_item_section(ws, item_name)
                
                cfg = load_app_config()
                stok_awal = float(cfg.get('stok_awal', {}).get(cfg.get('excel_filename', ''), {}).get(item_name, 0))
                stock_row = 5

                _write_stock_awal_row(ws, start_col, stock_row, stok_awal)
                if last_data_row < stock_row:
                    last_data_row = stock_row

                masuk_val, keluar_val = (float(qty), None) if masuk_keluar == "Masuk" else (None, float(qty))
                
                dup_row = check_duplicate(ws, col_header_row, tanggal, masuk_keluar, keterangan, start_col)
                write_row = dup_row if dup_row else last_data_row + 1
                
                ws.cell(row=write_row, column=start_col, value=tanggal)
                ws.cell(row=write_row, column=start_col + 1, value=masuk_val)
                ws.cell(row=write_row, column=start_col + 2, value=keluar_val)
                ws.cell(row=write_row, column=start_col + 3, value=None) # Akan diisi rumus oleh recalculate_item_stock
                ws.cell(row=write_row, column=start_col + 4, value=keterangan)
                
                # Sortir, warnai, dan masukkan rumus Excel
                recalculate_item_stock(ws, start_col, stock_row)
                
                written.append(f"{item_name} → {masuk_keluar} {qty} ({tanggal})")
            except Exception as e:
                errors.append(f"Error {item.get('name','?')}: {str(e)}")
    
    ok, msg = safe_save_workbook(wb, excel_path)
    if not ok:
        return jsonify({"status": "error", "message": msg, "written": written, "errors": errors})
    return jsonify({"status": "ok", "written": written, "errors": errors, "message": f"✅ {len(written)} baris ditulis"})

@stockin_bp.route('/api/sisa-stok', methods=['GET'])
def get_sisa_stok():
    excel_path, err = get_excel_path()
    if err:
        return jsonify({"status": "error", "message": err})
    if not os.path.exists(excel_path):
        return jsonify({"status": "ok", "sisa": {}})
    
    wb = openpyxl.load_workbook(excel_path)
    inv = load_inventory()
    sisa_map = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for item_cfg in inv['items']:
            if item_cfg.get('sheet') != sheet_name:
                continue
            try:
                _, last_data_row, last_sisa, _ = find_or_create_item_section(ws, item_cfg['display'])
                if last_data_row > 4:
                    sisa_map[item_cfg['id']] = last_sisa
            except:
                pass
    return jsonify({"status": "ok", "sisa": sisa_map})

@stockin_bp.route('/api/upload-pdf', methods=['POST'])
def upload_pdf():
    files = request.files.getlist('files')
    if not files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah"})

    inv = load_inventory()
    items_config = inv.get("items", [])
    
    cards = []
    
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            continue
        
        tmp_path = os.path.join('uploads', file.filename)
        os.makedirs('uploads', exist_ok=True)
        file.save(tmp_path)
        
        text = ""
        try:
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    extracted = page.extract_text()
                    if extracted:
                        text += extracted + "\n"
        except Exception as e:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            return jsonify({"status": "error", "message": f"Gagal membaca PDF: {str(e)}"})
            
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
            
        lines = text.split('\n')
        outlet = ""
        tanggal = ""
        
        # Coba parse dari nama file: "01 MAY 2026 003_GOKS!_GUNUNG_PANGILUN"
        name_no_ext = os.path.splitext(file.filename)[0]
        match_fn = re.match(r'^(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})\s+(.*)$', name_no_ext)
        if match_fn:
            day = match_fn.group(1).zfill(2)
            month_str = match_fn.group(2).upper()
            year = match_fn.group(3)
            rest_fn = match_fn.group(4)
            
            MONTHS = {
                "JANUARI": "01", "JAN": "01", "JANUARY": "01",
                "FEBRUARI": "02", "FEB": "02", "FEBRUARY": "02",
                "MARET": "03", "MAR": "03", "MARCH": "03",
                "APRIL": "04", "APR": "04",
                "MEI": "05", "MAY": "05",
                "JUNI": "06", "JUN": "06", "JUNE": "06",
                "JULI": "07", "JUL": "07", "JULY": "07",
                "AGUSTUS": "08", "AGU": "08", "AUG": "08", "AUGUST": "08",
                "SEPTEMBER": "09", "SEP": "09",
                "OKTOBER": "10", "OKT": "10", "OCT": "10", "OCTOBER": "10",
                "NOVEMBER": "11", "NOV": "11",
                "DESEMBER": "12", "DES": "12", "DEC": "12", "DECEMBER": "12"
            }
            if month_str in MONTHS:
                tanggal = f"{year}-{MONTHS[month_str]}-{day}"
                outlet = rest_fn.replace('_', ' ')
        
        for line in lines:
            line_lower = line.lower()
            if not outlet and ("divisi penerima" in line_lower or "kepada" in line_lower):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    outlet = parts[1].strip()
                else:
                    outlet = line_lower.replace("divisi penerima", "").replace("kepada", "").strip()
            
            if not tanggal and ("tanggal" in line_lower or "date" in line_lower):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    tanggal = parts[1].strip()
                else:
                    tanggal = line_lower.replace("tanggal", "").replace("date", "").strip()
                    
        extracted_items = []
        for line in lines:
            line_lower = line.lower()
            best_match_item = None
            best_match_kw = ""
            
            for item in items_config:
                for kw in item.get('pdf_keywords', []):
                    kw_lower = kw.lower()
                    # Gunakan regex agar pencocokan berdasarkan kata utuh, bukan potongan huruf
                    # (?<![a-z0-9]) memastikan sebelum keyword tidak ada huruf/angka
                    # (?![a-z0-9]) memastikan setelah keyword tidak ada huruf/angka
                    pattern = r'(?<![a-z0-9])' + re.escape(kw_lower) + r'(?![a-z0-9])'
                    if re.search(pattern, line_lower):
                        if len(kw_lower) > len(best_match_kw):
                            best_match_kw = kw_lower
                            best_match_item = item
            
            if best_match_item:
                item = best_match_item
                # Hapus keyword agar angka pada nama item tidak terbaca sebagai quantity
                line_without_kw = line_lower.replace(best_match_kw, '')
                nums = re.findall(r'\d+(?:[.,]\d+)?', line_without_kw)
                
                qty_raw = 0
                if nums:
                    try:
                        qty_str = nums[-1] if len(nums) > 0 else "0"
                        qty_raw = float(qty_str.replace(',', '.'))
                        if len(nums) > 1 and qty_raw < 10:
                            maybe_qty = float(nums[-2].replace(',', '.'))
                            if maybe_qty > qty_raw:
                                qty_raw = maybe_qty
                    except:
                        pass
                        
                pdf_unit_qty = item.get('pdf_unit_qty', 1)
                excel_factor = item.get('excel_factor', 1)
                qty_excel = (qty_raw / pdf_unit_qty) * excel_factor
                
                extracted_items.append({
                    "item_id": item['id'],
                    "name": item.get('display', item['id']),
                    "sheet": item.get('sheet', 'Bahan Lainnya'),
                    "qty_raw": qty_raw,
                    "qty_excel": qty_excel,
                    "unit": item.get('default_unit', ''),
                    "masuk_keluar": "Keluar"
                })
        
        cards.append({
            "original_filename": file.filename,
            "outlet": outlet,
            "tanggal": tanggal,
            "items": extracted_items
        })
        
    return jsonify({"status": "ok", "cards": cards})
