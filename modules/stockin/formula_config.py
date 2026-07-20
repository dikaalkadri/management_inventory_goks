"""
formula_config.py
=================
Konfigurasi dan penerapan rumus otomatis untuk sheet stok harian (01-31).

Untuk mengubah rumus:
  - Kolom J (fix per baris)  → edit J_FORMULAS
  - Kolom K, L, M (relative) → edit RELATIVE_FORMULAS / RELATIVE_ROW_START / RELATIVE_ROW_END
  - Rentang background J     → edit J_FILL_ROW_START / J_FILL_ROW_END
  - Kolom yang di-lock        → edit LOCKED_COLUMNS
  - Password proteksi         → edit SHEET_PASSWORD
"""

from openpyxl.styles import PatternFill, Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Protection, Border, Side

# ─── KONFIGURASI ─────────────────────────────────────────────────────────────

# Password proteksi sheet
SHEET_PASSWORD = "12345678"

# ── Rentang lock & kuning per kelompok kolom ──────────────────────────────────
# A(1), B(2)          → baris 1–95   (semua sheet)
# D(4)                → baris 1–95   (sheet 02-31 saja; sheet 01 TIDAK dikunci/kuning)
# I(9),J(10),K(11),
#   L(12),M(13)       → baris 2–95   (semua sheet)
# O(15),P(16),Q(17)   → baris 2–120  (semua sheet, dengan border)
# S(19),T(20),U(21)   → baris 2–19   (semua sheet)
# W(23),X(24)         → baris 1–123  (semua sheet; 122-123 = summary)

A_ROW_START    = 1;  A_ROW_END    = 95
D_ROW_START     = 1;  D_ROW_END     = 95
IJKLM_ROW_START = 2;  IJKLM_ROW_END = 95
OPQ_ROW_START   = 1;  OPQ_ROW_END   = 120
STU_ROW_START   = 1;  STU_ROW_END   = 19
WX_ROW_START    = 1;  WX_ROW_END    = 123

# Alias agar kompatibel dengan import di processor.py
RELATIVE_ROW_START = IJKLM_ROW_START
RELATIVE_ROW_END   = IJKLM_ROW_END

# Rentang background kuning kolom J (sama dengan IJKLM)
J_FILL_ROW_START = IJKLM_ROW_START
J_FILL_ROW_END   = IJKLM_ROW_END

# Baris summary W/X
WX_SUMMARY_START = 122   # W122 = "Selisih Bersih", X122 = SUM
WX_DATA_END      = 121   # baris terakhir data produk (untuk range SUM)



# Warna & border
FORMULA_FILL   = PatternFill("solid", fgColor="FFE6E6FA")   # kuning
HEADER_IJKLM_FILL = PatternFill("solid", fgColor="FFE6E6FA")
HEADER_WX_FILL = PatternFill("solid", fgColor="FFE6E6FA")   # oranye header W/X baris 1
BORDER_THIN    = Side(border_style="thin", color="000000")
CELL_BORDER    = Border(left=BORDER_THIN, right=BORDER_THIN,
                        top=BORDER_THIN,  bottom=BORDER_THIN)


# ─── RUMUS FIX KOLOM J ───────────────────────────────────────────────────────
# Key   = nomor baris (sesuai posisi di Excel, tidak berubah antar sheet/file)
# Value = rumus Excel persis seperti di file asli
# Untuk mengubah rumus → ganti value-nya
# Untuk mengubah posisi → ganti key-nya
# Untuk menambah rumus  → tambah baris baru
# Baris yang tidak ada di sini akan di-skip (tidak ditulis rumus, tapi tetap kuning)

import json
import os

# ── Cache in-memory untuk J Formulas ─────────────────────────────────────────
# Rumus kolom J sangat jarang berubah selama proses batch. Cache ini
# menghindari pembacaan file JSON dari disk sebanyak 1.000+ kali per batch.
_j_formulas_cache = None


def invalidate_formulas_cache():
    """Reset cache J Formulas. Dipanggil setelah rumus disimpan ulang."""
    global _j_formulas_cache
    _j_formulas_cache = None


def get_j_formulas():
    """Memuat rumus dari file JSON. Hasil di-cache di memori setelah baca pertama."""
    global _j_formulas_cache
    if _j_formulas_cache is not None:
        return _j_formulas_cache

    formulas = {}
    _json_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "mass_formulas.json")
    if os.path.exists(_json_path):
        try:
            with open(_json_path, "r", encoding="utf-8") as _f:
                _data = json.load(_f)
                for _k, _v in _data.items():
                    if str(_k).isdigit():
                        formulas[int(_k)] = _v.get("formula", "")
        except Exception as e:
            pass
    _j_formulas_cache = formulas
    return _j_formulas_cache


# ─── RUMUS RELATIVE KOLOM K, L, M ────────────────────────────────────────────
# {r} akan diganti dengan nomor baris saat penulisan
# Untuk mengubah rumus → ganti value-nya
# Untuk mengubah rentang baris → ubah RELATIVE_ROW_START / RELATIVE_ROW_END di atas

RELATIVE_FORMULAS = {
    "K": "=D{r}+E{r}-F{r}-G{r}",
    "L": "=D{r}+E{r}-J{r}",
    "M": "=J{r}-K{r}",
}

# ─── FUNGSI UTAMA ─────────────────────────────────────────────────────────────

def _safe_write(ws, row, col, value):
    """Tulis value ke cell, aman untuk MergedCell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    try:
        cell.value = value
    except:
        pass
    return cell


def _set_cell(ws, row, col, fill=None, locked=True, border=None, value=None):
    """
    Helper satu tempat: tulis value, set fill, border, dan lock sekaligus.
    Aman untuk MergedCell — cell merged di-skip.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    if value is not None:
        try:
            cell.value = value
        except:
            pass
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    cell.protection = Protection(locked=locked)


def apply_all(ws, lock_column_d=False):
    """
    Fungsi utama yang dipanggil dari processor.py.

    Arsitektur SATU PASS:
      1. Unlock semua cell (clean slate)
      2. Terapkan kuning + lock + rumus per kelompok kolom
      3. Aktifkan proteksi sheet
    Tidak ada fungsi yang saling menimpa karena semua dikerjakan
    dalam urutan linier tanpa unlock kedua.

    Kolom & rentang:
      A(1), B(2)        → kuning + lock, baris 1–95
      D(4)              → kuning + lock, baris 1–95, HANYA jika lock_column_d=True
      I(9)              → kuning + lock, baris 2–95  (nilai tetap, tidak ada rumus)
      J(10)             → kuning + lock + rumus J_FORMULAS, baris 2–95
      K(11),L(12),M(13) → kuning + lock + rumus relative, baris 2–95
      O(15),P(16),Q(17) → kuning + lock + border, baris 2–120
      S(19),T(20),U(21) → kuning + lock, baris 2–19
      W(23),X(24)       → kuning + lock + border + rumus, baris 1–123
        - W/X baris 1        : header oranye
        - W/X baris 2–121    : data produk (X = =W{r}*M{r})
        - W122               : label "Selisih Bersih", X122 = =SUM(X2:X121)
        - W123               : label "Kerugian Murni", X123 = =SUMIF(X2:X121,"<0")

    Parameter:
        lock_column_d (bool): True untuk sheet 02-31, False untuk sheet 01.
    """

    # ── STEP 1: Unlock semua cell (clean slate) ───────────────────────────────
    # Dibatasi sampai WX_ROW_END (baris 123) — baris tertinggi yang kita kelola.
    # Tidak perlu unlock baris kosong di bawahnya yang tidak pernah disentuh.
    # Penghematan: ~51% lebih sedikit operasi dibanding iter_rows() tanpa batas.
    for row in ws.iter_rows(max_row=WX_ROW_END):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.protection = Protection(locked=False)

    # ── STEP 2A: Kolom A (1) dan B (2) → baris 1–95, kuning + lock ───────────
    for col in [1]:
        for r in range(A_ROW_START, A_ROW_END + 1):
            _set_cell(ws, r, col, fill=FORMULA_FILL, locked=True)

    # ── STEP 2B: Kolom D (4) → baris 1–95, kuning + lock (sheet 02-31 saja) ──
    if lock_column_d:
        for r in range(D_ROW_START, D_ROW_END + 1):
            _set_cell(ws, r, 4, fill=FORMULA_FILL, locked=True)

    # ── HEADER I-M (baris 1) ────────────────────────────────────────────────
    for col in [9, 10, 11, 12, 13]:
        _set_cell(
            ws,
            1,
            col,
            fill=HEADER_IJKLM_FILL,
            locked=True
        )

    # ── STEP 2C: Kolom I (9) → baris 2–95, kuning + lock ────────────────────
    for r in range(IJKLM_ROW_START, IJKLM_ROW_END + 1):
        _set_cell(ws, r, 9, fill=FORMULA_FILL, locked=True)

    # ── STEP 2D: Kolom J (10) → baris 2–95, kuning + lock + rumus ───────────
    for r in range(J_FILL_ROW_START, J_FILL_ROW_END + 1):
        _set_cell(ws, r, 10, fill=FORMULA_FILL, locked=True)
        
    j_formulas = get_j_formulas()
    for row_num, formula in j_formulas.items():
        _safe_write(ws, row_num, 10, formula)

    # ── STEP 2E: Kolom K(11), L(12), M(13) → baris 2–95, kuning+lock+rumus ──
    col_map = {"K": 11, "L": 12, "M": 13}
    for col_letter, col_idx in col_map.items():
        for r in range(RELATIVE_ROW_START, RELATIVE_ROW_END + 1):
            formula = RELATIVE_FORMULAS[col_letter].format(r=r)
            _set_cell(ws, r, col_idx, fill=FORMULA_FILL, locked=True, value=formula)

    # ── STEP 2F: Kolom O(15), P(16), Q(17) → baris 2–120, kuning+lock+border─
    # Bersihkan dulu label lama di baris 118 & 119 kolom W jika ada
    for r in range(OPQ_ROW_START, OPQ_ROW_END + 1):
        for col in [15, 16, 17]:
            _set_cell(ws, r, col, fill=FORMULA_FILL, locked=True, border=CELL_BORDER)

    # ── STEP 2G: Kolom S(19), T(20), U(21) → baris 2–19, kuning + lock ──────
    for col in [19, 20, 21]:
        for r in range(STU_ROW_START, STU_ROW_END + 1):
            _set_cell(ws, r, col, fill=FORMULA_FILL, locked=True)

    # ── STEP 2H: Kolom W(23) dan X(24) → baris 1–122 ────────────────────────
    # Bersihkan label lama "Selisih Bersih" / "Kerugian Murni" di baris 118-119
    for r in [118, 119]:
        cell_w = ws.cell(row=r, column=23)
        if not isinstance(cell_w, MergedCell):
            old_val = str(cell_w.value or "")
            if "Selisih" in old_val or "Kerugian" in old_val:
                cell_w.value = None

    # Baris 1: header oranye
    for col in [23, 24]:
        _set_cell(ws, 1, col, fill=HEADER_WX_FILL, locked=True, border=CELL_BORDER)

    # Baris 2–120: data produk, kuning, rumus X = W*M
    for r in range(2, WX_DATA_END + 1):
        _set_cell(ws, r, 23, fill=FORMULA_FILL, locked=True, border=CELL_BORDER)
        _set_cell(ws, r, 24, fill=FORMULA_FILL, locked=True, border=CELL_BORDER,
                  value=f"=W{r}*M{r}")

    # Baris 122: Selisih Bersih
    _set_cell(ws, WX_SUMMARY_START, 23, fill=FORMULA_FILL, locked=True,
              border=CELL_BORDER, value="Selisih Bersih")
    _set_cell(ws, WX_SUMMARY_START, 24, fill=FORMULA_FILL, locked=True,
              border=CELL_BORDER, value=f"=SUM(X2:X{WX_DATA_END})")

    # Baris 123: Kerugian Murni
    _set_cell(ws, WX_SUMMARY_START + 1, 23, fill=FORMULA_FILL, locked=True,
              border=CELL_BORDER, value="Kerugian Murni")
    _set_cell(ws, WX_SUMMARY_START + 1, 24, fill=FORMULA_FILL, locked=True,
              border=CELL_BORDER, value=f'=SUMIF(X2:X{WX_DATA_END},"<0")')

    # ── STEP 3: Aktifkan proteksi sheet ──────────────────────────────────────
    ws.protection.sheet = True
    ws.protection.password = SHEET_PASSWORD
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False   # izinkan resize kolom
    ws.protection.formatRows = False      # izinkan resize baris