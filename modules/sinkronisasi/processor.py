"""
Core logic for Sales & Loss Synchronization (SO Kerugian).
Handles raw sales data extraction, workbook rows synchronization, category auto-merging,
formulas and protection application, conditional formatting, and summary sheet generation.
"""
import os
import re
import calendar as _cal
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from services.catalog_service import load_pos, load_materials
from services.excel_style_helper import copy_row_properties
from modules.stockin.formula_config import apply_all as apply_formulas_and_protect, RELATIVE_ROW_START, RELATIVE_ROW_END
from modules.stockin.helpers import load_inventory, safe_save_workbook

def safe_write_cell(ws, row, col, value, number_format=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row:
                if col >= merged_range.min_col and col <= merged_range.max_col:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
    try:
        cell.value = value
        if number_format:
            cell.number_format = number_format
    except:
        pass

def proses_sinkronisasi_excel(sumber_path_xlsx, tujuan_path_xlsx, output_folder, outlet=None):
    pos_list = load_pos()
    mat_list = load_materials()
    total_items = max(len(pos_list), len(mat_list))
    if total_items == 0:
        raise ValueError("Data POS Menu dan Bahan Baku kosong.")

    # ── TAHAP 1: EKSTRAK DATA PENJUALAN SUMBER ──
    wb_sumber = openpyxl.load_workbook(sumber_path_xlsx, read_only=True, data_only=True)
    ws_sumber = wb_sumber.worksheets[0]

    sales_map = {}
    available_days = set()
    def clean_key(text):
        return re.sub(r'[^A-Z0-9]', '', str(text or "").upper())

    for r, row in enumerate(ws_sumber.iter_rows(values_only=True), start=1):
        if r < 2:
            continue
        if not row or len(row) < 17:
            continue
        
        date_val = row[0]
        if date_val is None:
            continue
        
        prod_val = row[15]
        qty_val = row[16]
        if not prod_val:
            continue

        day_str = None
        if isinstance(date_val, (datetime, date)):
            day_str = f"{date_val.day:02d}"
        else:
            match = re.search(r'^(\d{1,2})[\/\-]', str(date_val).strip())
            if match:
                day_str = f"{int(match.group(1)):02d}"

        if not day_str:
            continue
            
        available_days.add(day_str)
        
        try:
            qty = float(qty_val)
        except:
            qty = 0
        
        key = (day_str, clean_key(prod_val))
        sales_map[key] = sales_map.get(key, 0) + qty

    wb_sumber.close()

    # ── TAHAP 2: PROSES WORKBOOK TUJUAN ──
    wb_tujuan = openpyxl.load_workbook(tujuan_path_xlsx)
    updated_sheets = []
    
    tujuan_filename = os.path.basename(tujuan_path_xlsx)
    bulan_str, tahun_str = "mei", "2026"
    bulan_list = ['januari','februari','maret','april','mei','juni','juli','agustus','september','oktober','november','desember']
    for m_name in bulan_list:
        if m_name in tujuan_filename.lower():
            bulan_str = m_name
            break
    bulan_num = bulan_list.index(bulan_str) + 1
    tahun_int = int(tahun_str)

    for sheet_name in wb_tujuan.sheetnames:
        try:
            day_num = int(sheet_name.strip())
            if day_num < 1 or day_num > 31:
                continue
            day_str = f"{day_num:02d}"
        except ValueError:
            continue
            
        should_update_sales = (day_str in available_days)

        ws_tujuan = wb_tujuan[sheet_name]
        ws_tujuan.freeze_panes = 'C2'

        # Update header tanggal di kolom Q (17) row 1 sesuai tanggal sheet
        try:
            sheet_date = datetime(tahun_int, bulan_num, day_num)
            date_cell = ws_tujuan.cell(row=1, column=17)
            date_cell.value = sheet_date
            date_cell.number_format = 'dd/mm/yyyy'
        except ValueError:
            pass  # Skip jika tanggal tidak valid (misal tgl 31 di bulan 30 hari)

        # Tulis header kolom W dan X jika belum ada
        if not ws_tujuan.cell(row=1, column=23).value:
            ws_tujuan.cell(row=1, column=23).value = "Harga Produk"
        if not ws_tujuan.cell(row=1, column=24).value:
            ws_tujuan.cell(row=1, column=24).value = "Nilai Kerugian"

        # 1. Cari titik baris header "PRODUCT"
        start_data_row = None
        for r in range(1, 25):
            val_p = str(ws_tujuan.cell(row=r, column=16).value or "").strip().upper()
            if "PRODUCT" in val_p:
                start_data_row = r + 1
                break
        if not start_data_row:
            start_data_row = 2

        # 2. Cari batas bawah area data (sebelum "TOTAL")
        current_row_idx = start_data_row
        while True:
            val_p = ws_tujuan.cell(row=current_row_idx, column=16).value
            if val_p and any(keyword in str(val_p).upper() for keyword in ["TOTAL", "JUMLAH"]):
                break
            if not val_p and not ws_tujuan.cell(row=current_row_idx + 1, column=16).value:
                break
            current_row_idx += 1
            
        old_last_data_row = current_row_idx - 1
        existing_row_count = (old_last_data_row - start_data_row) + 1
        target_last_data_row = start_data_row + total_items - 1

        # 3. UNMERGE SEL KATEGORI LAMA DI KOLOM O (15) SECARA DINAMIS
        merges_to_remove = []
        for merged_range in list(ws_tujuan.merged_cells.ranges):
            # Cari merge yang berada di dalam area data lama dan khusus untuk Kolom Kategori (15)
            if merged_range.min_row >= start_data_row and merged_range.max_row <= old_last_data_row:
                if merged_range.min_col == 15 and merged_range.max_col == 15:
                    merges_to_remove.append(merged_range)
        
        # Hapus proteksi read-only pada kategori
        for m in merges_to_remove:
            try:
                ws_tujuan.unmerge_cells(str(m))
            except:
                pass

        # 4. Tambah / Hapus baris sesuai selisih dengan Master Catalog
        if existing_row_count < total_items:
            rows_to_add = total_items - existing_row_count
            ws_tujuan.insert_rows(old_last_data_row + 1, amount=rows_to_add)
            for r_new in range(old_last_data_row + 1, target_last_data_row + 1):
                copy_row_properties(ws_tujuan, start_data_row, r_new)
        elif existing_row_count > total_items:
            rows_to_delete = existing_row_count - total_items
            ws_tujuan.delete_rows(target_last_data_row + 1, amount=rows_to_delete)

        # 5. Tulis Data per baris (Karena Kolom O sudah di-unmerge, akan 100% aman)
        for idx in range(total_items):
            r = start_data_row + idx
            
            # --- Tulis Data POS (Kolom O, P, Q) ---
            if idx < len(pos_list):
                item = pos_list[idx]
                p_name = item.get('product', '').strip()
                cat = item.get('category', '')
                
                safe_write_cell(ws_tujuan, r, 15, cat)            # Kolom O (Category POS)
                safe_write_cell(ws_tujuan, r, 16, p_name)         # Kolom P (Product POS)

                p_key = clean_key(p_name)
                
                if should_update_sales:
                    sales_qty = sales_map.get((day_str, p_key), 0)
                    safe_write_cell(ws_tujuan, r, 17, sales_qty)      # Kolom Q (Paid Qty)
            else:
                safe_write_cell(ws_tujuan, r, 15, "")
                safe_write_cell(ws_tujuan, r, 16, "")
                if should_update_sales:
                    safe_write_cell(ws_tujuan, r, 17, "")

            # --- Tulis Data Bahan Baku (Kolom A, B, W) ---
            if idx < len(mat_list):
                mat = mat_list[idx]
                mat_name = mat.get('name', '').strip()
                mat_cat = mat.get('category', '')
                price = float(mat.get('price', 0))
                
                safe_write_cell(ws_tujuan, r, 1, mat_cat)         # Kolom A (Kategori Bahan)
                safe_write_cell(ws_tujuan, r, 2, mat_name)        # Kolom B (Nama Bahan)
                safe_write_cell(ws_tujuan, r, 23, price, '#,##0') # Kolom W (Harga Modal)
            else:
                safe_write_cell(ws_tujuan, r, 1, "")
                safe_write_cell(ws_tujuan, r, 2, "")
                safe_write_cell(ws_tujuan, r, 23, "")

        # 5a. Berikan warna ZigZag (Zebra) pada baris data sebelum rumus diterapkan
        COLOR_EVEN_BG = "F5F7FA" # Warna biru-abu sangat muda dan soft
        COLOR_ODD_BG  = "FFFFFF"
        for row_idx in range(start_data_row, target_last_data_row + 1):
            is_even = (row_idx - start_data_row) % 2 == 0
            fill_color = COLOR_EVEN_BG if is_even else COLOR_ODD_BG
            pattern = PatternFill("solid", fgColor=fill_color)
            for col_idx in range(1, 14): # Kolom A (1) sampai M (13) saja
                cell = ws_tujuan.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.fill = pattern

        # 5b. Rumus Stock Awal kolom D dari Stok Akhir (kolom F) sheet sebelumnya
        #     Sheet "01" → input manual, kolom D tidak dikunci/kuning
        #     Sheet "02"-"31" → D{r} = ='{prev}'!F{r}
        is_first_sheet = (day_num == 1)
        if not is_first_sheet:
            prev_day_str = f"{day_num - 1:02d}"
            if prev_day_str in wb_tujuan.sheetnames:
                for r in range(RELATIVE_ROW_START, RELATIVE_ROW_END + 1):
                    cell_d = ws_tujuan.cell(row=r, column=4)
                    if not isinstance(cell_d, MergedCell):
                        cell_d.value = f"='{prev_day_str}'!F{r}"

        # 5c. Terapkan semua kuning + lock + rumus + proteksi (satu pass)
        apply_formulas_and_protect(ws_tujuan, lock_column_d=not is_first_sheet)

        # 6. RE-MERGE KOLOM KATEGORI BERDASARKAN SUSUNAN BARU (Untuk POS Menu)
        current_cat = None
        start_cat_row = start_data_row
        
        for idx in range(len(pos_list)):
            r = start_data_row + idx
            cat = pos_list[idx].get('category', '').strip()
            
            if current_cat is None:
                current_cat = cat
                start_cat_row = r
            elif current_cat != cat:
                # Gabungkan baris kategori yang sama (jika lebih dari 1 baris)
                if r - 1 > start_cat_row:
                    ws_tujuan.merge_cells(start_row=start_cat_row, start_column=15, end_row=r - 1, end_column=15)
                    ws_tujuan.cell(row=start_cat_row, column=15).alignment = Alignment(horizontal='center', vertical='center')
                
                current_cat = cat
                start_cat_row = r
        
        # Eksekusi merge untuk blok kategori paling akhir
        if len(pos_list) > 0:
            last_r = start_data_row + len(pos_list) - 1
            if last_r > start_cat_row:
                ws_tujuan.merge_cells(start_row=start_cat_row, start_column=15, end_row=last_r, end_column=15)
                ws_tujuan.cell(row=start_cat_row, column=15).alignment = Alignment(horizontal='center', vertical='center')
            
        # 6b. BERI GARIS TEBAL BAWAH UNTUK KATEGORI BAHAN BAKU (Kolom A sampai M)
        thick_side = Side(style='medium', color='000000')
        current_mat_cat = None

        for idx in range(len(mat_list)):
            r = start_data_row + idx
            mat_cat = mat_list[idx].get('category', '').strip()
            
            if current_mat_cat is None:
                current_mat_cat = mat_cat
            elif current_mat_cat != mat_cat:
                # Beri garis tebal bawah untuk kategori sebelumnya (Kolom A sampai M)
                for col_idx in range(1, 14):
                    cell = ws_tujuan.cell(row=r - 1, column=col_idx)
                    cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=thick_side)
                
                current_mat_cat = mat_cat
        
        # Garis tebal untuk kategori bahan baku terakhir
        if len(mat_list) > 0:
            last_mat_r = start_data_row + len(mat_list) - 1
            for col_idx in range(1, 14):
                cell = ws_tujuan.cell(row=last_mat_r, column=col_idx)
                cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=thick_side)

        # 7. Conditional Formatting: semua sel < 0 di area data → background merah
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        red_font = Font(color="CC0000", bold=True)
        red_rule = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)

        # Terapkan ke seluruh area data (semua kolom, semua baris data)
        data_range = f"A{start_data_row}:Z{target_last_data_row}"
        ws_tujuan.conditional_formatting.add(data_range, red_rule)

        updated_sheets.append(sheet_name)

    # ── TAHAP 3: BUAT / UPDATE SHEET "Rekap Kerugian" ──
    REKAP_SHEET = "Rekap Kerugian"

    # Hapus sheet lama jika sudah ada (replace)
    if REKAP_SHEET in wb_tujuan.sheetnames:
        del wb_tujuan[REKAP_SHEET]

    # Buat sheet baru, letakkan di posisi pertama
    ws_rekap = wb_tujuan.create_sheet(REKAP_SHEET, 0)

    COLOR_HEADER_BG   = "1B5E20"   # hijau gelap
    COLOR_HEADER_FONT = "FFFFFF"
    COLOR_TOTAL_BG    = "F9FBE7"
    COLOR_EVEN_BG     = "F1F8E9"
    COLOR_ODD_BG      = "FFFFFF"
    COLOR_ZERO_FONT   = "BDBDBD"   # abu-abu untuk tgl yang tidak ada datanya

    thin = Side(border_style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Segoe UI", bold=True, color=COLOR_HEADER_FONT, size=11)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def data_cell(ws, row, col, value, fmt=None, even=True, zero=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Segoe UI", size=10, color=COLOR_ZERO_FONT if zero else "000000")
        c.fill = PatternFill("solid", fgColor=COLOR_EVEN_BG if even else COLOR_ODD_BG)
        c.alignment = Alignment(horizontal="center" if fmt is None else "right", vertical="center")
        c.border = border
        if fmt:
            c.number_format = fmt
        return c

    # ── Judul ──
    ws_rekap.row_dimensions[1].height = 30
    title_cell = ws_rekap.cell(row=1, column=1, value=f"REKAP KERUGIAN — {bulan_str.upper()} {tahun_str}")
    title_cell.font = Font(name="Segoe UI", bold=True, size=14, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_rekap.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # ── Header kolom ──
    ws_rekap.row_dimensions[2].height = 36
    hdr_cell(ws_rekap, 2, 1, "No",                              width=6)
    hdr_cell(ws_rekap, 2, 2, "Tanggal",                        width=18)
    hdr_cell(ws_rekap, 2, 3, "Selisih Bersih (Surplus+Defisit)", width=28)
    hdr_cell(ws_rekap, 2, 4, "Kerugian Murni (Defisit Saja)",    width=28)
    hdr_cell(ws_rekap, 2, 5, "Keterangan",                     width=24)

    # ── Data baris tgl 01 s.d. 31 ──
    bulan_num_local = bulan_list.index(bulan_str) + 1
    _, max_days_in_month = _cal.monthrange(tahun_int, bulan_num_local)

    rp_fmt = '_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"_);_(@_)'

    for day in range(1, 32):
        row = day + 2   # data mulai baris 3
        even = (day % 2 == 0)
        sheet_exists = f"{day:02d}" in wb_tujuan.sheetnames and day <= max_days_in_month

        # Nomor
        data_cell(ws_rekap, row, 1, day, even=even, zero=not sheet_exists)

        # Tanggal
        try:
            tgl = datetime(tahun_int, bulan_num_local, day) if day <= max_days_in_month else None
        except ValueError:
            tgl = None
        c_tgl = data_cell(ws_rekap, row, 2, tgl, even=even, zero=(tgl is None))
        if tgl:
            c_tgl.number_format = "dd/mm/yyyy"
            c_tgl.font = Font(name="Segoe UI", size=10)

        if sheet_exists:
            # Rumus INDIRECT — ROW(A1) di baris rekap hari ke-1, ROW(A2) hari ke-2, dst
            # Hasil di Excel: =INDIRECT("'"&TEXT(ROW(A1),"00")&"'!X123")
            a_ref = f"A{day}"
            f_bersih = '=INDIRECT("\'"&TEXT(ROW(' + a_ref + '),"00")&"\'!X122")'
            f_murni  = '=INDIRECT("\'"&TEXT(ROW(' + a_ref + '),"00")&"\'!X123")'
            data_cell(ws_rekap, row, 3, f_bersih, fmt=rp_fmt, even=even)
            data_cell(ws_rekap, row, 4, f_murni,  fmt=rp_fmt, even=even)
        else:
            # Tgl tidak ada di workbook atau melebihi hari dalam bulan → nilai 0
            data_cell(ws_rekap, row, 3, 0, fmt=rp_fmt, even=even, zero=True)
            data_cell(ws_rekap, row, 4, 0, fmt=rp_fmt, even=even, zero=True)

        data_cell(ws_rekap, row, 5, "", even=even)

    # ── Baris TOTAL ──
    total_row = 34   # baris 3 + 31 hari = 34
    ws_rekap.row_dimensions[total_row].height = 22
    for col in range(1, 6):
        c = ws_rekap.cell(row=total_row, column=col)
        c.font = Font(name="Segoe UI", bold=True, size=11, color=COLOR_HEADER_FONT)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws_rekap.cell(row=total_row, column=1).value = ""
    ws_rekap.cell(row=total_row, column=2).value = "TOTAL"

    c_tot_bersih = ws_rekap.cell(row=total_row, column=3, value=f"=SUM(C3:C33)")
    c_tot_bersih.number_format = rp_fmt
    c_tot_bersih.font = Font(name="Segoe UI", bold=True, color=COLOR_HEADER_FONT, size=11)
    c_tot_bersih.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c_tot_bersih.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_bersih.border = border

    c_tot_murni = ws_rekap.cell(row=total_row, column=4, value=f"=SUM(D3:D33)")
    c_tot_murni.number_format = rp_fmt
    c_tot_murni.font = Font(name="Segoe UI", bold=True, color=COLOR_HEADER_FONT, size=11)
    c_tot_murni.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c_tot_murni.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_murni.border = border

    # Freeze header
    ws_rekap.freeze_panes = "A3"

    # Penamaan Output File: Otomatis berdasarkan nama file
    sumber_filename = os.path.basename(sumber_path_xlsx) if sumber_path_xlsx else "unknown"
    full_outlet_string = "unknown"
    try:
        # load_inventory sudah di-import di awal file
        inv = load_inventory()
        outlets_list = inv.get("outlets", [])
        
        tujuan_check = tujuan_filename.replace('_', ' ').lower()
        sumber_check = sumber_filename.replace('_', ' ').lower()
        
        # Prioritaskan outlet dari parameter jika ada (dari dropdown fallback)
        if outlet and outlet.strip():
            full_outlet_string = outlet.strip()
        else:
            for o in outlets_list:
                o_lower = o.lower()
                if o_lower in tujuan_check or o_lower in sumber_check:
                    full_outlet_string = o
                    break
                # Fallback: cari namanya saja tanpa nomor (misal "taplau")
                o_name = re.sub(r'^\d+\.\s*', '', o).lower()
                if o_name and (o_name in tujuan_check or o_name in sumber_check):
                    full_outlet_string = o
                    break
    except Exception as e:
        print(f"[DEBUG SINKRONISASI] Error saat deteksi nama outlet: {e}")

    clean_outlet_name = re.sub(r'[^a-z0-9]', '_', full_outlet_string.lower()).strip('_')
    output_filename = f"kerugian_{clean_outlet_name}_{bulan_str}_{tahun_str}.xlsx"
    
    output_filepath = os.path.join(output_folder, output_filename)

    ok, msg = safe_save_workbook(wb_tujuan, output_filepath)
    if not ok:
        raise IOError(msg)
        
    return output_filename, len(updated_sheets)