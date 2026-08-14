"""
processor.py – Add Stock Auth
===============================
Membaca file SO Drive lalu menghasilkan 1 file Excel
berisi rekap data di sheet pertama (Add Stock Auth) yang ditambahkan ke workbook SO Drive.

Sesuai requirement:
- Menggunakan 29 Master Outlet.
- Menghasilkan rumus =SUM() untuk Outlet dan Gudang.
- Mendeteksi nama outlet yang salah dan menaruhnya di kolom WRONG NAME.
"""

import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from services.task_manager import active_tasks

VALID_OUTLETS = [
    "AIRDINGIN", "AIRPACAH", "AIRTAWAR", "ANDURING", "BALAIBARU", "BANDARBUAT", 
    "BELIBIS", "BELIMBING", "BUNGUS", "CENDRAWASIH", "GP", "JATI", "KANTOR", 
    "KASANG", "KURANJI", "KURAITAJI", "LUBAY", "LUBEG", "LUBUKALUNG", "LUBUKLINTAH", 
    "PARAKLAWEH", "PASBAR", "RAWANG", "SEBPDG", "SITEBA", "TABING", "TAPLAU", 
    "TC", "TH"
]

def _norm(text: str) -> str:
    if not text or (isinstance(text, float)): return ""
    return re.sub(r'[^a-z0-9]', '', str(text).lower())

def _scan_so_drive(so_path: str):
    xls = pd.ExcelFile(so_path, engine='openpyxl')
    available_sheets = xls.sheet_names
    so_data = {}
    
    for sheet_name in available_sheets:
        df = pd.read_excel(so_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        if df.empty: continue
            
        n_cols = len(df.columns)
        col_offset = 0
        
        while col_offset < n_cols:
            header_val = df.iloc[0, col_offset]
            if pd.isna(header_val):
                col_offset += 6
                continue
                
            item_name = str(header_val).strip()
            if not item_name:
                col_offset += 6
                continue
                
            norm_key = _norm(item_name)
            if not norm_key:
                col_offset += 6
                continue
                
            keluar_col_letter = get_column_letter(col_offset + 3)
            
            if norm_key not in so_data:
                so_data[norm_key] = {
                    'outlets': {},
                    'wrong_names': set(),
                    'sheet': sheet_name,
                    'keluar_col': keluar_col_letter,
                    'original_name': item_name
                }
            
            data_rows = df.iloc[2:, col_offset:col_offset+5]
            for _, row in data_rows.iterrows():
                if len(row) < 5: continue
                ket_val = row.iloc[4]
                keluar_val = row.iloc[2]
                
                if pd.isna(ket_val) or pd.isna(keluar_val): continue
                    
                ket_str = str(ket_val).strip().upper()
                if not ket_str.startswith('U/'): continue
                    
                outlet_code = ket_str[2:].strip()
                
                try: qty = float(keluar_val)
                except: continue
                    
                if qty <= 0: continue
                
                if outlet_code in VALID_OUTLETS:
                    so_data[norm_key]['outlets'][outlet_code] = so_data[norm_key]['outlets'].get(outlet_code, 0.0) + qty
                else:
                    so_data[norm_key]['wrong_names'].add(outlet_code)
                    
            col_offset += 6
            
    return so_data

def process_add_stock_auth(task_id: str, so_bytes: bytes, output_folder: str) -> dict:
    import tempfile
    active_tasks[task_id] = {'progress': 0, 'total': 100, 'current_item': 'Menyimpan file sementara...', 'status': 'running'}
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_so:
            tmp_so.write(so_bytes)
            so_path = tmp_so.name
            
        active_tasks[task_id] = {'progress': 30, 'total': 100, 'current_item': 'Scanning SO Drive...', 'status': 'running'}
        so_data = _scan_so_drive(so_path)
        
        if not so_data:
            raise Exception("Data kosong atau format salah. Pastikan Anda mengunggah File SO Drive (bukan file Template).")
        
        active_tasks[task_id] = {'progress': 60, 'total': 100, 'current_item': 'Membuat file output Excel...', 'status': 'running'}
        wb = openpyxl.load_workbook(so_path)
        
        ws = wb.create_sheet("Add Stock Auth", 0)
        wb.active = ws
        
        # ── Color palette ──────────────────────────────────────────────
        PURPLE       = "5A189A"   # header bg (item name col)
        BLUE_HDR     = "1D3557"   # header bg (outlet cols)
        GREEN_HDR    = "1B4332"   # header bg (SUM per outlet)
        ORANGE_HDR   = "7B3F00"   # header bg (STOCKOUT)
        VIOLET_HDR   = "3D1A78"   # header bg (SELISIH)
        RED_HDR      = "7B1D1D"   # header bg (WRONG NAME)
        WHITE        = "FFFFFF"

        # Row fills – alternating
        ROW_EVEN     = "EEF2FF"   # light lavender
        ROW_ODD      = "FFFFFF"   # white
        SUM_EVEN     = "D1FAE5"   # light green
        SUM_ODD      = "ECFDF5"
        STOCK_EVEN   = "FEF3C7"   # light amber
        STOCK_ODD    = "FFFBEB"
        SEL_EVEN     = "EDE9FE"   # light violet
        SEL_ODD      = "F5F3FF"
        WRONG_FILL   = "FEE2E2"   # light red (always)

        def hfill(hex_str):
            return PatternFill(start_color=hex_str, end_color=hex_str, fill_type="solid")
        def rfill(hex_str):
            return PatternFill(start_color=hex_str, end_color=hex_str, fill_type="solid") if hex_str else None

        font_header_white  = Font(bold=True, color=WHITE, size=10)
        font_header_yellow = Font(bold=True, color="FFD700", size=10)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left   = Alignment(horizontal="left", vertical="center")
        thin = Side(style='thin', color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Freeze first row and first column
        ws.freeze_panes = "B2"

        # ── Header row height ──────────────────────────────────────────
        ws.row_dimensions[1].height = 32

        # Col A – "Nama Produk/ Outlet"
        c = ws.cell(row=1, column=1, value="Nama Produk/ Outlet")
        c.fill  = hfill(PURPLE)
        c.font  = font_header_white
        c.alignment = align_left
        c.border = border
        ws.column_dimensions['A'].width = 32

        # Outlet columns
        col_idx = 2
        for outlet in VALID_OUTLETS:
            cell = ws.cell(row=1, column=col_idx, value=f"U/{outlet}")
            cell.fill      = hfill(BLUE_HDR)
            cell.font      = font_header_white
            cell.alignment = align_center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
            col_idx += 1

        # SUM PER OUTLET
        col_sum_outlet = col_idx
        c = ws.cell(row=1, column=col_sum_outlet, value="SUM PER OUTLET")
        c.fill = hfill(GREEN_HDR); c.font = font_header_white
        c.alignment = align_center; c.border = border
        ws.column_dimensions[get_column_letter(col_sum_outlet)].width = 18

        # STOCKOUT GUDANG
        col_stockout = col_idx + 1
        c = ws.cell(row=1, column=col_stockout, value="STOCKOUT GUDANG")
        c.fill = hfill(ORANGE_HDR); c.font = font_header_white
        c.alignment = align_center; c.border = border
        ws.column_dimensions[get_column_letter(col_stockout)].width = 20

        # SELISIH
        col_selisih = col_idx + 2
        c = ws.cell(row=1, column=col_selisih, value="SELISIH")
        c.fill = hfill(VIOLET_HDR); c.font = font_header_white
        c.alignment = align_center; c.border = border
        ws.column_dimensions[get_column_letter(col_selisih)].width = 13

        # WRONG NAME
        col_wrong = col_idx + 3
        c = ws.cell(row=1, column=col_wrong, value="WRONG NAME")
        c.fill = hfill(RED_HDR); c.font = font_header_yellow
        c.alignment = align_center; c.border = border
        ws.column_dimensions[get_column_letter(col_wrong)].width = 28

        row_idx = 2
        for norm_key, found_data in so_data.items():
            is_even = (row_idx % 2 == 0)
            ws.row_dimensions[row_idx].height = 18

            # Col A – item name
            c = ws.cell(row=row_idx, column=1, value=found_data['original_name'])
            c.fill = hfill("E0D7F5" if is_even else "F3EEFF")  # soft purple zebra
            c.font = Font(bold=True, size=10, color="3D0066")
            c.alignment = align_left
            c.border = border

            # Outlet qty columns
            c_idx = 2
            row_bg = ROW_EVEN if is_even else ROW_ODD
            for outlet in VALID_OUTLETS:
                qty = found_data['outlets'].get(outlet, 0)
                cell = ws.cell(row=row_idx, column=c_idx, value=qty)
                cell.fill = hfill(row_bg)
                cell.alignment = align_center
                cell.border = border
                if qty == 0:
                    cell.font = Font(color="AAAAAA", size=9)  # dim zeros
                else:
                    cell.font = Font(color="1D3557", bold=True, size=9)
                c_idx += 1

            # SUM PER OUTLET
            outlet_start = get_column_letter(2)
            outlet_end   = get_column_letter(col_sum_outlet - 1)
            sum_formula  = f"=SUM({outlet_start}{row_idx}:{outlet_end}{row_idx})"
            cell = ws.cell(row=row_idx, column=col_sum_outlet, value=sum_formula)
            cell.fill = hfill(SUM_EVEN if is_even else SUM_ODD)
            cell.font = Font(bold=True, color="1B4332", size=10)
            cell.alignment = align_center; cell.border = border

            # STOCKOUT GUDANG
            sheet       = found_data['sheet']
            keluar_col  = found_data['keluar_col']
            stock_formula = f"=SUM('{sheet}'!{keluar_col}:{keluar_col})"
            cell = ws.cell(row=row_idx, column=col_stockout, value=stock_formula)
            cell.fill = hfill(STOCK_EVEN if is_even else STOCK_ODD)
            cell.font = Font(bold=True, color="7B3F00", size=10)
            cell.alignment = align_center; cell.border = border

            # SELISIH
            sum_letter     = get_column_letter(col_sum_outlet)
            stockout_letter = get_column_letter(col_stockout)
            sel_formula    = f"={stockout_letter}{row_idx}-{sum_letter}{row_idx}"
            cell = ws.cell(row=row_idx, column=col_selisih, value=sel_formula)
            cell.fill = hfill(SEL_EVEN if is_even else SEL_ODD)
            cell.font = Font(bold=True, color="3D1A78", size=10)
            cell.alignment = align_center; cell.border = border

            # WRONG NAME
            wrong_name_val = ""
            if found_data['wrong_names']:
                wrong_name_val = ", ".join(sorted(list(found_data['wrong_names'])))
            cell = ws.cell(row=row_idx, column=col_wrong, value=wrong_name_val)
            if wrong_name_val:
                cell.fill = hfill(WRONG_FILL)
                cell.font = Font(bold=True, color="991B1B", size=9)
            else:
                cell.fill = hfill("F9FAFB" if is_even else WHITE)
                cell.font = Font(color="9CA3AF", size=9)
            cell.alignment = align_left; cell.border = border

            row_idx += 1
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Add_Stock_Auth_{timestamp}.xlsx"
        output_path = os.path.join(output_folder, filename)
        
        active_tasks[task_id] = {'progress': 90, 'total': 100, 'current_item': 'Menyimpan file hasil...', 'status': 'running'}
        wb.save(output_path)
        
        os.remove(so_path)
        
        active_tasks[task_id] = {'progress': 100, 'total': 100, 'current_item': 'Selesai', 'status': 'completed',
            'result': {
                'filename': filename,
                'download_url': f"/api/add-stock-auth/download/{filename}",
                'total_items': len(so_data)
            }
        }
        return active_tasks[task_id]['result']
        
    except Exception as e:
        active_tasks[task_id] = {'progress': 100, 'total': 100, 'current_item': f'Error: {str(e)}', 'status': 'error', 'message': str(e)}
        return {'error': str(e)}
